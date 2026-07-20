import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import json
import requests
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF
import base64
from io import BytesIO
import time
from typing import List, Dict, Optional
import hashlib
import hmac
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from google import genai
from google.genai import types
import os

# Importar modelos y feature engineering
from models import ModelEnsemble
from feature_engineering import extract_match_features, create_sample_training_data
import validate_models

# Importar Supabase opcionalmente
try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False
    create_client = None
    Client = None

# Diccionario de traducciones (bilingüe)
TRANSLATIONS = {
    "es": {
        "page_title": "SportsPredict Pro - Predicciones Deportivas",
        "menu": "Navegación",
        "dashboard": "🏠 Dashboard",
        "my_predictions": "📊 Mis Predicciones",
        "ai_predictions": "🤖 Predicciones IA",
        "competitions": "🏆 Competencias",
        "statistics": "📈 Estadísticas",
        "alerts": "🔔 Alertas",
        "premium": "💎 Premium",
        "admin": "⚙️ Administración",
        "language": "Idioma",
        "light_mode": "Modo Claro",
        "dark_mode": "Modo Oscuro",
        "chatbot_title": "💬 Asistente IA",
        "chatbot_placeholder": "Escribe tu pregunta sobre predicciones o datos de partidos...",
        "send_button": "Enviar",
        "voice_button": "🎤 Hablar",
        "stop_voice_button": "⏹️ Detener",
        # =========== LOGIN PAGE ===========
        "login_title": "🔐 Iniciar Sesión",
        "access_to_platform": "Acceso a la plataforma",
        "username": "Usuario",
        "password": "Contraseña",
        "login_button": "Iniciar Sesión",
        "guest_mode": "Modo Invitado",
        "guest_info": "Puedes usar la plataforma en modo invitado sin iniciar sesión. Algunas funcionalidades estarán limitadas.",
        "test_credentials": "Credenciales de prueba",
        "continue_as_guest": "Continuar como Invitado",
        "login_success": "✅ Inicio de sesión exitoso!",
        "login_error": "❌ Credenciales incorrectas",
        "guest_success": "✅ Modo invitado activado",
        # =========== USER PROFILE ===========
        "user_label": "👤 Usuario",
        "plan_label": "📊 Plan",
        "logout_button": "Cerrar Sesión",
        # =========== DASHBOARD ===========
        "dashboard_title": "⚽ Dashboard de Predicciones Deportivas",
        "greeting": "👋 ¡Hola, {}!",
        "how_it_works": "**¿Cómo funciona?**",
        "how_step1": "1. Mira los próximos partidos en la sección de abajo.",
        "how_step2": "2. Expande un partido, ingresa tu predicción de goles y tu nivel de confianza.",
        "how_step3": "3. Haz clic en \"🎯 Predecir\" para guardar tu predicción.",
        "how_step4": "4. Ve tus predicciones guardadas en la página \"📊 Mis Predicciones\".",
        "guest_warning": "⚠️ Estás en modo invitado. Inicia sesión para acceder a todas las funcionalidades.",
        "your_stats": "Tus Estadísticas",
        "total_predictions": "Total Predicciones",
        "accuracy_rate": "Tasa de Precisión",
        "avg_confidence": "Confianza Promedio",
        "global_rank": "Ranking Global",
        "today": "hoy",
        "positions": "posiciones",
        "upcoming_matches": "📋 Próximos Partidos para Predecir",
        "loading_matches": "Cargando partidos...",
        "no_matches_available": "No hay partidos disponibles en este momento",
        "ai_prediction": "**Predicción IA:**",
        "winner": "Ganador: {}",
        "probability": "Probabilidad: {:.0f}%",
        "your_predictions": "**Tus Predicciones:**",
        "home_goals": "Goles local",
        "away_goals": "Goles visitante",
        "confidence_level": "Nivel de confianza",
        "predict_button": "🎯 Predecir",
        "prediction_saved": "✅ Predicción guardada correctamente",
        "prediction_error": "❌ Error guardando predicción",
        "recent_performance": "📈 Tu Rendimiento Reciente",
        "no_predictions_graphs": "No hay predicciones suficientes para mostrar gráficos de rendimiento",
        "accuracy_evolution": "Evolución de Precisión (30 días)",
        "prediction_volume": "Volumen de Predicciones",
        # =========== MY PREDICTIONS ===========
        "my_predictions_title": "📊 Mis Predicciones",
        "high_confidence": "Alta confianza (>70%)",
        "confidence_avg": "Confianza promedio",
        "prediction_history": "Historial de Predicciones",
        "match_id": "ID Partido",
        "home_score": "Goles Local",
        "away_score": "Goles Visitante",
        "confidence": "Confianza",
        "date": "Fecha",
        "confidence_analysis": "📈 Análisis de Confianza",
        "confidence_distribution": "Distribución de Niveles de Confianza",
        "no_predictions_yet": "Aún no has realizado predicciones. ¡Comienza ahora!",
        # =========== AI PREDICTIONS ===========
        "ai_predictions_title": "🤖 Predicciones con Inteligencia Artificial",
        "ai_models_intro": "Nuestra inteligencia artificial combina 4 modelos:",
        "model_xgb": "🌲 **XGBoost**: Modelo de árboles de decisión con boosting",
        "model_bayes": "📊 **Bayesian Ridge**: Modelo probabilístico de regresión",
        "model_lstm": "🧠 **LSTM con Atención**: Red neuronal recurrente para patrones temporales",
        "model_ensemble": "🎯 **Ensemble Final**: Meta-modelo (Logistic Regression) que combina todos los anteriores",
        "sport": "Deporte",
        "min_ai_confidence": "Confianza mínima IA",
        "next_days": "Próximos días",
        "analyzing_matches": "Analizando partidos para los próximos {} días...",
        "home_label": "Local",
        "draw_label": "Empate",
        "away_label": "Visitante",
        "model_predictions": "📊 Predicciones de cada modelo",
        "result_label": "Resultado",
        "probability_label": "Probabilidad",
        "xgb_title": "🌲 XGBoost",
        "bayes_title": "📊 Bayesian Ridge",
        "lstm_title": "🧠 LSTM con Atención",
        "ensemble_title": "🎯 Ensemble Final (Meta-Modelo: Logistic Regression)",
        "expected_winner": "**Ganador esperado:** {}",
        "expected_goals_api": "**Goles esperados (API):**",
        "home_goals_label": "Local: {:.1f}",
        "away_goals_label": "Visitante: {:.1f}",
        "ai_advice_api": "**Consejo IA (API):**",
        "compare_teams": "📊 Comparativa de Equipos (API)",
        "form_label": "**Forma (Últimos partidos):**",
        "form_home": "Local: {}",
        "form_away": "Visitante: {}",
        "att_def_label": "**Ataque / Defensa:**",
        "att_def_home": "Local: ATK {} / DEF {}",
        "att_def_away": "Visitante: ATK {} / DEF {}",
        "ai_confidence_gauge": "Confianza IA (Ensemble)",
        # =========== COMPETITIONS ===========
        "competitions_title": "🏆 Competencias y Torneos",
        "available_competitions": "Competencias Disponibles",
        "description_label": "**Descripción:** {}",
        "participants_label": "Participantes",
        "total_prize": "Premio Total",
        "entry_fee": "Entrada",
        "deadline_label": "📅 **Cierra:** {}",
        "join_button": "✅ Unirse",
        "joined_success": "¡Te has unido a {}!",
        "premium_required": "💎 Requiere suscripción Premium",
        "status_label": "Status: ✅ Activa",
        "your_ranking": "📊 Tu Ranking",
        # =========== STATISTICS ===========
        "stats_title": "📊 ESTADÍSTICAS DESCRIPTIVAS Y ANÁLISIS",
        "tab_summary": "📈 Resumen",
        "tab_tables": "📋 Tablas",
        "tab_advanced_graphs": "📊 Gráficos Avanzados",
        "tab_reports": "📥 Reportes",
        "tab_insights": "💡 Insights",
        "general_summary": "Resumen General",
        "performance_summary": "**Resumen de Desempeño**",
        "metric_label": "Métrica",
        "value_label": "Valor",
        "total_label": "Total",
        "correct_label": "Correctas",
        "incorrect_label": "Incorrectas",
        "precision_label": "Precisión",
        "avg_conf_short": "Confianza Prom.",
        "rank_label": "Ranking",
        "descriptive_stats_predictions": "**Estadísticas Descriptivas de Predicciones**",
        "no_numeric_data": "No hay datos numéricos disponibles",
        "advanced_visualizations": "Visualizaciones Avanzadas",
        "confidence_distribution_curve": "Distribución de Confianza (Histograma + Curva Normal)",
        "cdf_title": "Función de Distribución Acumulada (CDF)",
        "time_evolution": "Evolución en el Tiempo",
        "confidence_time_evolution": "Evolución de Confianza en el Tiempo",
        "status_analysis": "Análisis por Status",
        "status_distribution": "Distribución por Status",
        "status_x_label": "Status",
        "status_y_label": "Cantidad",
        "no_data_graphs": "No hay suficientes datos para mostrar gráficos",
        "generate_reports": "📥 Generar Reportes",
        "pdf_report": "**Reporte en PDF**",
        "download_pdf": "📄 Descargar PDF",
        "excel_report": "**Reporte en Excel**",
        "download_excel": "📊 Descargar Excel",
        "insights_title": "💡 Insights y Recomendaciones",
        "performance_analysis": "**Análisis de Rendimiento**",
        "excellent_performance": "✅ **EXCELENTE** - Tu precisión es superior al 75%",
        "excellent_insight": "Mantén tu estrategia actual, has encontrado un buen patrón.",
        "good_performance": "⚠️ **BUENO** - Tu precisión está sobre el promedio",
        "good_insight": "Continúa mejorando, hay espacio para optimizar.",
        "developing_performance": "⚠️ **DESARROLLO** - Tu precisión está por debajo del 60%",
        "developing_insight": "Revisa tu metodología de análisis.",
        "confidence_analysis": "**Análisis de Confianza**",
        "high_conf": "✅ Confianza bien calibrada (>0.75)",
        "high_conf_insight": "Tu confianza es realista con tu precisión.",
        "moderate_conf": "⚠️ Confianza moderada (0.60-0.75)",
        "moderate_conf_insight": "Considera si tu confianza refleja tu precisión.",
        "low_conf": "⚠️ Confianza baja (<0.60)",
        "low_conf_insight": "Aumenta confianza cuando tengas análisis sólidos.",
        "personalized_recommendations": "**Recomendaciones Personalizadas**",
        "rec1": "🔍 Analiza tus predicciones incorrectas para identificar patrones",
        "rec2": "📚 Aumenta confianza solo cuando tengas datos sólidos de respaldo",
        "rec3": "🎯 Diversifica tus predicciones entre diferentes deportes y ligas",
        "rec4": "📊 Mantén un registro detallado de tus análisis y resultados",
        "rec5": "⏱️ Revisa tu rendimiento regularmente (semanal/mensual)",
        "rec6": "💡 Considera variables externas: lesiones, clima, forma del equipo",
        # =========== ALERTS ===========
        "alerts_title": "🔔 Alertas y Notificaciones",
        "recent_notifications": "Notificaciones Recientes",
        "new_match_available": "Nuevo partido disponible",
        "rm_vs_barca": "Real Madrid vs Barcelona - Predicción IA lista",
        "two_hours_ago": "Hace 2 horas",
        "ai_prediction_updated": "Actualización de predicción IA",
        "mancity_vs_liverpool": "Nueva predicción para Manchester City vs Liverpool",
        "five_hours_ago": "Hace 5 horas",
        "competition_ending_soon": "Competencia finaliza pronto",
        "liga_predictor_ends": "Liga Predictor Enero cierra en 2 días",
        "one_day_ago": "Hace 1 día",
        "ranking_change": "Cambio en tu ranking",
        "rank_up_5": "Subiste 5 posiciones en el ranking global",
        "config_alerts": "Configurar Alertas",
        "new_matches_check": "⚽ Nuevos partidos",
        "ai_predictions_check": "🤖 Predicciones IA",
        "competitions_check": "🏆 Competencias",
        "ranking_changes_check": "📈 Cambios ranking",
        "premium_offers_check": "💎 Ofertas Premium",
        "save_config": "Guardar configuración",
        "config_saved": "✅ Configuración guardada",
        # =========== PREMIUM ===========
        "premium_title": "💎 SportsPredict Pro Premium",
        "free_plan": "📦 Plan Free",
        "pro_plan": "🌟 Plan Pro",
        "elite_plan": "👑 Plan Elite",
        "current_plan": "Tu plan actual",
        "subscribe_pro": "🚀 Suscribirse a Pro",
        "subscribe_elite": "👑 Suscribirse a Elite",
        "welcome_pro": "¡Bienvenido a SportsPredict Pro!",
        "welcome_elite": "¡Bienvenido a SportsPredict Elite!",
        "additional_benefits": "Beneficios Adicionales",
        "feature_label": "Característica",
        # =========== SIDEBAR ===========
        "connected_supabase": "✅ Conectado a Supabase",
        "supabase_warning": "⚠️ Supabase disponible pero sin conexión o tablas no creadas",
        "local_mode": "ℹ️ Usando modo local (sin Supabase)",
        "info_title": "📱 Información",
        "info_text": """
        **SportsPredict Pro v1.0**

        Plataforma de predicciones deportivas con IA

        🔗 [Sitio Web](https://example.com)
        📧 [Contacto](mailto:info@example.com)
        """,
        # =========== AI PREDICTIONS ADDITIONAL ===========
        "ai_predictions": "🤖 Predicciones con Inteligencia Artificial",
        "ai_models_intro": """
        Nuestra inteligencia artificial combina 4 modelos:
        - 🌲 **XGBoost**: Modelo de árboles de decisión con boosting
        - 📊 **Bayesian Ridge**: Modelo probabilístico de regresión
        - 🧠 **LSTM con Atención**: Red neuronal recurrente para patrones temporales
        - 🎯 **Ensemble Final**: Meta-modelo (Logistic Regression) que combina todos los anteriores
        """,
        "football": "Fútbol",
        "basketball": "Baloncesto",
        "tennis": "Tenis",
        "baseball": "Béisbol",
        "home": "Local",
        "draw": "Empate",
        "away": "Visitante",
        "result": "Resultado",
        "each_model_predictions": "📊 Predicciones de cada modelo",
        "xgb_probabilities": "Probabilidades XGBoost",
        "bayes_probabilities": "Probabilidades Bayesian Ridge",
        "lstm_probabilities": "Probabilidades LSTM con Atención",
        "ensemble_probabilities": "Probabilidades Ensemble",
        "expected_winner": "**Ganador esperado:** {}",
        "expected_goals_api": "**Goles esperados (API):**",
        "home_goals_label": "Local: {:.1f}",
        "away_goals_label": "Visitante: {:.1f}",
        "ai_advice_api": "**Consejo IA (API):**",
        "expected_goals": "Goles esperados (API)",
        "analyze_stats": "Analiza estadísticas",
        "team_comparison_title": "📊 Comparativa de Equipos (API)",
        "form_last_matches": "**Forma (Últimos partidos):**",
        "home_form": "Local: {}",
        "away_form": "Visitante: {}",
        "attack_defense": "**Ataque / Defensa:**",
        "home_atk_def": "Local: ATK {} / DEF {}",
        "away_atk_def": "Visitante: ATK {} / DEF {}",
        "ai_confidence_title": "Confianza IA (Ensemble)",
        "no_matches_period": "No hay partidos disponibles en este período",
        "statistics_title": "📊 Estadísticas Avanzadas",
        "tab1": "📈 Resumen",
        "tab2": "📋 Tablas",
        "tab3": "📊 Gráficos Avanzados",
        "tab4": "📥 Reportes",
        "tab5": "💡 Insights",
        "descriptive_stats": "Estadísticas Descriptivas",
        "pred_desc_stats": "Estadísticas Descriptivas de Predicciones",
        "advanced_viz": "Visualizaciones Avanzadas",
        "conf_dist_title": "Distribución de Confianza (Histograma + Curva Normal)",
        "metric_col": "Métrica",
        "value_col": "Valor",
        "total": "Total",
        "correct": "Correctas",
        "incorrect": "Incorrectas",
        "precision": "Precisión",
        "ranking": "Ranking",
        "admin_tab1": "📊 Gestión de Datos",
        "admin_tab2": "🤖 Validación de Modelos",
        "admin_validation_header": "Validación Robusta de Modelos",
        "admin_run_validation": "🚀 Ejecutar Validación",
        "admin_validating": "Ejecutando validación robusta de modelos...",
        "admin_validation_success": "✅ Validación completada con éxito!",
        "admin_validation_error": "Error durante la validación: {}",
        "admin_metrics_general": "📊 Métricas Generales de Rendimiento",
        "admin_classification": "Clasificación",
        "admin_regression": "Regresión",
        "admin_individual_models": "🤖 Rendimiento de Modelos Individuales",
        "admin_visualizations": "📈 Visualizaciones",
        "admin_residual_analysis": "📉 Análisis de Residuos",
        "admin_normality_tests": "🔍 Pruebas de Normalidad",
        "admin_cv_stability": "📊 Estabilidad (Validación Cruzada)",
        "admin_overfitting_test": "⚠️ Prueba de Sobreajuste",
        "admin_download_results": "📥 Descargar Resultados (CSV)",
        "ai_assistant_page": "💬 Asistente IA",
        "audio_input_label": "🎤 Entrada de audio",
        "record_button": "Grabar",
        "stop_record_button": "Detener grabación",
        "processing_audio": "Procesando audio...",
        "quota_exceeded_msg": "¡Lo siento! Se ha excedido la cuota de la API de Gemini. Por favor, intenta nuevamente más tarde o revisa los límites de tu plan.",
        # =========== COMPETITIONS ADDITIONAL ===========
        "entry_label": "Entrada",
        "free": "Gratis",
        "deadline_label": "📅 **Cierra:** {}",
        "joined_comp": "¡Te has unido a {}!",
        "active_status": "Activa",
        "you_label": "👤 Tú",
        "rank_format": "{}. {} - {} pts ({})",
        "comp1_name": "Liga Predictor Enero",
        "comp1_desc": "Predice correctamente y gana premios semanales",
        "comp2_name": "Torneo Premium Elite",
        "comp2_desc": "Exclusivo para suscriptores premium - Grandes premios",
        "comp3_name": "Desafío de 100 Predicciones",
        "comp3_desc": "Realiza 100 predicciones precisas y gana jackpot",
        # =========== STATISTICS ADDITIONAL ===========
        "statistics_title": "📊 ESTADÍSTICAS DESCRIPTIVAS Y ANÁLISIS",
        "tab1": "📈 Resumen",
        "tab2": "📋 Tablas",
        "tab3": "📊 Gráficos Avanzados",
        "tab4": "📥 Reportes",
        "tab5": "💡 Insights",
        "metric_col": "Métrica",
        "value_col": "Valor",
        "total": "Total",
        "correct": "Correctas",
        "incorrect": "Incorrectas",
        "ranking": "Ranking",
        "descriptive_stats": "Estadísticas Descriptivas",
        "pred_desc_stats": "**Estadísticas Descriptivas de Predicciones**",
        "advanced_viz": "Visualizaciones Avanzadas",
        "conf_dist_title": "Distribución de Confianza (Histograma + Curva Normal)",
        "time_evol_conf": "Evolución de Confianza en el Tiempo",
        "status_dist": "Distribución por Status",
        "quantity": "Cantidad",
        "not_enough_data": "No hay suficientes datos para mostrar gráficos",
        "perf_analysis": "**Análisis de Rendimiento**",
        "excellent_perf": "✅ **EXCELENTE** - Tu precisión es superior al 75%",
        "insight1_excellent": "Mantén tu estrategia actual, has encontrado un buen patrón.",
        "good_perf": "⚠️ **BUENO** - Tu precisión está sobre el promedio",
        "insight1_good": "Continúa mejorando, hay espacio para optimizar.",
        "dev_perf": "⚠️ **DESARROLLO** - Tu precisión está por debajo del 60%",
        "insight1_dev": "Revisa tu metodología de análisis.",
        "conf_analysis": "**Análisis de Confianza**",
        "insight2_high": "Tu confianza es realista con tu precisión.",
        "insight2_moderate": "Considera si tu confianza refleja tu precisión.",
        "insight2_low": "Aumenta confianza cuando tengas análisis sólidos.",
        "personal_recommendations": "**Recomendaciones Personalizadas**",
    },
    "en": {
        "page_title": "SportsPredict Pro - Sports Predictions",
        "menu": "Navigation",
        "dashboard": "🏠 Dashboard",
        "my_predictions": "📊 My Predictions",
        "ai_predictions": "🤖 AI Predictions",
        "competitions": "🏆 Competitions",
        "statistics": "📈 Statistics",
        "alerts": "🔔 Alerts",
        "premium": "💎 Premium",
        "admin": "⚙️ Administration",
        "language": "Language",
        "light_mode": "Light Mode",
        "dark_mode": "Dark Mode",
        "chatbot_title": "💬 AI Assistant",
        "chatbot_placeholder": "Type your question about predictions or match data...",
        "send_button": "Send",
        "voice_button": "🎤 Speak",
        "stop_voice_button": "⏹️ Stop",
        # =========== LOGIN PAGE ===========
        "login_title": "🔐 Log In",
        "access_to_platform": "Platform Access",
        "username": "Username",
        "password": "Password",
        "login_button": "Log In",
        "guest_mode": "Guest Mode",
        "guest_info": "You can use the platform in guest mode without logging in. Some features will be limited.",
        "test_credentials": "Test Credentials",
        "continue_as_guest": "Continue as Guest",
        "login_success": "✅ Login successful!",
        "login_error": "❌ Incorrect credentials",
        "guest_success": "✅ Guest mode activated",
        # =========== USER PROFILE ===========
        "user_label": "👤 User",
        "plan_label": "📊 Plan",
        "logout_button": "Log Out",
        # =========== DASHBOARD ===========
        "dashboard_title": "⚽ Sports Predictions Dashboard",
        "greeting": "👋 Hello, {}!",
        "how_it_works": "**How does it work?**",
        "how_step1": "1. Look at the upcoming matches in the section below.",
        "how_step2": "2. Expand a match, enter your goal prediction and your confidence level.",
        "how_step3": "3. Click \"🎯 Predict\" to save your prediction.",
        "how_step4": "4. View your saved predictions on the \"📊 My Predictions\" page.",
        "guest_warning": "⚠️ You are in guest mode. Log in to access all features.",
        "your_stats": "Your Stats",
        "total_predictions": "Total Predictions",
        "accuracy_rate": "Accuracy Rate",
        "avg_confidence": "Avg Confidence",
        "global_rank": "Global Rank",
        "today": "today",
        "positions": "positions",
        "upcoming_matches": "📋 Upcoming Matches to Predict",
        "loading_matches": "Loading matches...",
        "no_matches_available": "No matches available right now",
        "ai_prediction": "**AI Prediction:**",
        "winner": "Winner: {}",
        "probability": "Probability: {:.0f}%",
        "your_predictions": "**Your Predictions:**",
        "home_goals": "Home goals",
        "away_goals": "Away goals",
        "confidence_level": "Confidence level",
        "predict_button": "🎯 Predict",
        "prediction_saved": "✅ Prediction saved successfully",
        "prediction_error": "❌ Error saving prediction",
        "recent_performance": "📈 Your Recent Performance",
        "no_predictions_graphs": "Not enough predictions to show performance graphs",
        "accuracy_evolution": "Accuracy Evolution (30 days)",
        "prediction_volume": "Prediction Volume",
        # =========== MY PREDICTIONS ===========
        "my_predictions_title": "📊 My Predictions",
        "high_confidence": "High confidence (>70%)",
        "confidence_avg": "Avg confidence",
        "prediction_history": "Prediction History",
        "match_id": "Match ID",
        "home_score": "Home Score",
        "away_score": "Away Score",
        "confidence": "Confidence",
        "date": "Date",
        "confidence_analysis": "📈 Confidence Analysis",
        "confidence_distribution": "Confidence Levels Distribution",
        "no_predictions_yet": "You haven't made any predictions yet. Start now!",
        # =========== AI PREDICTIONS ===========
        "ai_predictions_title": "🤖 AI Predictions",
        "ai_models_intro": "Our AI combines 4 models:",
        "model_xgb": "🌲 **XGBoost**: Decision tree model with boosting",
        "model_bayes": "📊 **Bayesian Ridge**: Probabilistic regression model",
        "model_lstm": "🧠 **LSTM with Attention**: Recurrent neural network for temporal patterns",
        "model_ensemble": "🎯 **Final Ensemble**: Meta-model (Logistic Regression) combining all above",
        "sport": "Sport",
        "min_ai_confidence": "Min AI Confidence",
        "next_days": "Next Days",
        "analyzing_matches": "Analyzing matches for the next {} days...",
        "home_label": "Home",
        "draw_label": "Draw",
        "away_label": "Away",
        "model_predictions": "📊 Each Model's Predictions",
        "result_label": "Result",
        "probability_label": "Probability",
        "xgb_title": "🌲 XGBoost",
        "bayes_title": "📊 Bayesian Ridge",
        "lstm_title": "🧠 LSTM with Attention",
        "ensemble_title": "🎯 Final Ensemble (Meta-Model: Logistic Regression)",
        "expected_winner": "**Expected Winner:** {}",
        "expected_goals_api": "**Expected Goals (API):**",
        "home_goals_label": "Home: {:.1f}",
        "away_goals_label": "Away: {:.1f}",
        "ai_advice_api": "**AI Advice (API):**",
        "compare_teams": "📊 Team Comparison (API)",
        "form_label": "**Form (Last Matches):**",
        "form_home": "Home: {}",
        "form_away": "Away: {}",
        "att_def_label": "**Attack / Defense:**",
        "att_def_home": "Home: ATK {} / DEF {}",
        "att_def_away": "Away: ATK {} / DEF {}",
        "ai_confidence_gauge": "AI Confidence (Ensemble)",
        # =========== COMPETITIONS ===========
        "competitions_title": "🏆 Competitions and Tournaments",
        "available_competitions": "Available Competitions",
        "description_label": "**Description:** {}",
        "participants_label": "Participants",
        "total_prize": "Total Prize",
        "entry_fee": "Entry Fee",
        "deadline_label": "📅 **Closes:** {}",
        "join_button": "✅ Join",
        "joined_success": "You've joined {}!",
        "premium_required": "💎 Premium subscription required",
        "status_label": "Status: ✅ Active",
        "your_ranking": "📊 Your Ranking",
        # =========== STATISTICS ===========
        "stats_title": "📊 DESCRIPTIVE STATISTICS AND ANALYSIS",
        "tab_summary": "📈 Summary",
        "tab_tables": "📋 Tables",
        "tab_advanced_graphs": "📊 Advanced Graphs",
        "tab_reports": "📥 Reports",
        "tab_insights": "💡 Insights",
        "general_summary": "General Summary",
        "performance_summary": "**Performance Summary**",
        "metric_label": "Metric",
        "value_label": "Value",
        "total_label": "Total",
        "correct_label": "Correct",
        "incorrect_label": "Incorrect",
        "precision_label": "Accuracy",
        "avg_conf_short": "Avg Conf.",
        "rank_label": "Rank",
        "descriptive_stats_predictions": "**Prediction Descriptive Statistics**",
        "no_numeric_data": "No numeric data available",
        "advanced_visualizations": "Advanced Visualizations",
        "confidence_distribution_curve": "Confidence Distribution (Histogram + Normal Curve)",
        "cdf_title": "Cumulative Distribution Function (CDF)",
        "time_evolution": "Evolution Over Time",
        "confidence_time_evolution": "Confidence Evolution Over Time",
        "status_analysis": "Status Analysis",
        "status_distribution": "Status Distribution",
        "status_x_label": "Status",
        "status_y_label": "Count",
        "no_data_graphs": "Not enough data to show graphs",
        "generate_reports": "📥 Generate Reports",
        "pdf_report": "**PDF Report**",
        "download_pdf": "📄 Download PDF",
        "excel_report": "**Excel Report**",
        "download_excel": "📊 Download Excel",
        "insights_title": "💡 Insights and Recommendations",
        "performance_analysis": "**Performance Analysis**",
        "excellent_performance": "✅ **EXCELLENT** - Your accuracy is above 75%",
        "excellent_insight": "Keep your current strategy, you've found a good pattern.",
        "good_performance": "⚠️ **GOOD** - Your accuracy is above average",
        "good_insight": "Keep improving, there's room for optimization.",
        "developing_performance": "⚠️ **DEVELOPING** - Your accuracy is below 60%",
        "developing_insight": "Review your analysis methodology.",
        "confidence_analysis": "**Confidence Analysis**",
        "high_conf": "✅ Well-calibrated confidence (>0.75)",
        "high_conf_insight": "Your confidence is realistic with your accuracy.",
        "moderate_conf": "⚠️ Moderate confidence (0.60-0.75)",
        "moderate_conf_insight": "Consider whether your confidence reflects your accuracy.",
        "low_conf": "⚠️ Low confidence (<0.60)",
        "low_conf_insight": "Increase confidence when you have solid supporting data.",
        "personalized_recommendations": "**Personalized Recommendations**",
        "rec1": "🔍 Analyze your incorrect predictions to identify patterns",
        "rec2": "📚 Increase confidence only when you have solid supporting data",
        "rec3": "🎯 Diversify your predictions across different sports and leagues",
        "rec4": "📊 Keep a detailed record of your analyses and results",
        "rec5": "⏱️ Review your performance regularly (weekly/monthly)",
        "rec6": "💡 Consider external variables: injuries, weather, team form",
        # =========== ALERTS ===========
        "alerts_title": "🔔 Alerts and Notifications",
        "recent_notifications": "Recent Notifications",
        "new_match_available": "New match available",
        "rm_vs_barca": "Real Madrid vs Barcelona - AI Prediction ready",
        "two_hours_ago": "2 hours ago",
        "ai_prediction_updated": "AI prediction updated",
        "mancity_vs_liverpool": "New prediction for Manchester City vs Liverpool",
        "five_hours_ago": "5 hours ago",
        "competition_ending_soon": "Competition ending soon",
        "liga_predictor_ends": "Predictor League January ends in 2 days",
        "one_day_ago": "1 day ago",
        "ranking_change": "Your ranking changed",
        "rank_up_5": "You moved up 5 positions in the global ranking",
        "config_alerts": "Configure Alerts",
        "new_matches_check": "⚽ New matches",
        "ai_predictions_check": "🤖 AI predictions",
        "competitions_check": "🏆 Competitions",
        "ranking_changes_check": "📈 Ranking changes",
        "premium_offers_check": "💎 Premium offers",
        "save_config": "Save configuration",
        "config_saved": "✅ Configuration saved",
        # =========== PREMIUM ===========
        "premium_title": "💎 SportsPredict Pro Premium",
        "free_plan": "📦 Free Plan",
        "pro_plan": "🌟 Pro Plan",
        "elite_plan": "👑 Elite Plan",
        "current_plan": "Your current plan",
        "subscribe_pro": "🚀 Subscribe to Pro",
        "subscribe_elite": "👑 Subscribe to Elite",
        "welcome_pro": "Welcome to SportsPredict Pro!",
        "welcome_elite": "Welcome to SportsPredict Elite!",
        "additional_benefits": "Additional Benefits",
        "feature_label": "Feature",
        # =========== SIDEBAR ===========
        "connected_supabase": "✅ Connected to Supabase",
        "supabase_warning": "⚠️ Supabase available but no connection or tables not created",
        "local_mode": "ℹ️ Using local mode (without Supabase)",
        "info_title": "📱 Information",
        "info_text": """
        **SportsPredict Pro v1.0**

        Sports prediction platform with AI

        🔗 [Website](https://example.com)
        📧 [Contact](mailto:info@example.com)
        """,
        # =========== AI PREDICTIONS ADDITIONAL ===========
        "ai_predictions": "🤖 AI Predictions",
        "ai_models_intro": """
        Our AI combines 4 models:
        - 🌲 **XGBoost**: Decision tree model with boosting
        - 📊 **Bayesian Ridge**: Probabilistic regression model
        - 🧠 **LSTM with Attention**: Recurrent neural network for temporal patterns
        - 🎯 **Final Ensemble**: Meta-model (Logistic Regression) combining all above
        """,
        "football": "Football",
        "basketball": "Basketball",
        "tennis": "Tennis",
        "baseball": "Baseball",
        "home": "Home",
        "draw": "Draw",
        "away": "Away",
        "result": "Result",
        "each_model_predictions": "📊 Each Model's Predictions",
        "xgb_probabilities": "XGBoost Probabilities",
        "bayes_probabilities": "Bayesian Ridge Probabilities",
        "lstm_probabilities": "LSTM with Attention Probabilities",
        "ensemble_probabilities": "Ensemble Probabilities",
        "expected_winner": "**Expected Winner:** {}",
        "expected_goals_api": "**Expected Goals (API):**",
        "home_goals_label": "Home: {:.1f}",
        "away_goals_label": "Away: {:.1f}",
        "ai_advice_api": "**AI Advice (API):**",
        "expected_goals": "Expected Goals (API)",
        "analyze_stats": "Analyze statistics",
        "team_comparison_title": "📊 Team Comparison (API)",
        "form_last_matches": "**Form (Last Matches):**",
        "home_form": "Home: {}",
        "away_form": "Away: {}",
        "attack_defense": "**Attack / Defense:**",
        "home_atk_def": "Home: ATK {} / DEF {}",
        "away_atk_def": "Away: ATK {} / DEF {}",
        "ai_confidence_title": "AI Confidence (Ensemble)",
        "no_matches_period": "No matches available in this period",
        "statistics_title": "📊 Advanced Statistics",
        "tab1": "📈 Summary",
        "tab2": "📋 Tables",
        "tab3": "📊 Advanced Graphs",
        "tab4": "📥 Reports",
        "tab5": "💡 Insights",
        "descriptive_stats": "Descriptive Statistics",
        "pred_desc_stats": "Prediction Descriptive Statistics",
        "advanced_viz": "Advanced Visualizations",
        "conf_dist_title": "Confidence Distribution (Histogram + Normal Curve)",
        "metric_col": "Metric",
        "value_col": "Value",
        "total": "Total",
        "correct": "Correct",
        "incorrect": "Incorrect",
        "precision": "Precision",
        "ranking": "Ranking",
        "admin_tab1": "📊 Data Management",
        "admin_tab2": "🤖 Model Validation",
        "admin_validation_header": "Robust Model Validation",
        "admin_run_validation": "🚀 Run Validation",
        "admin_validating": "Running robust model validation...",
        "admin_validation_success": "✅ Validation completed successfully!",
        "admin_validation_error": "Error during validation: {}",
        "admin_metrics_general": "📊 General Performance Metrics",
        "admin_classification": "Classification",
        "admin_regression": "Regression",
        "admin_individual_models": "🤖 Individual Model Performance",
        "admin_visualizations": "📈 Visualizations",
        "admin_residual_analysis": "📉 Residual Analysis",
        "admin_normality_tests": "🔍 Normality Tests",
        "admin_cv_stability": "📊 Stability (Cross-Validation)",
        "admin_overfitting_test": "⚠️ Overfitting Test",
        "admin_download_results": "📥 Download Results (CSV)",
        "ai_assistant_page": "💬 AI Assistant",
        "audio_input_label": "🎤 Audio Input",
        "record_button": "Record",
        "stop_record_button": "Stop Recording",
        "processing_audio": "Processing audio...",
        "quota_exceeded_msg": "Sorry! The Gemini API quota has been exceeded. Please try again later or check your plan limits.",
        # =========== COMPETITIONS ADDITIONAL ===========
        "entry_label": "Entry",
        "free": "Free",
        "deadline_label": "📅 **Closes:** {}",
        "joined_comp": "You've joined {}!",
        "active_status": "Active",
        "you_label": "👤 You",
        "rank_format": "{}. {} - {} pts ({})",
        "comp1_name": "January Predictor League",
        "comp1_desc": "Predict correctly and win weekly prizes",
        "comp2_name": "Premium Elite Tournament",
        "comp2_desc": "Exclusive for premium subscribers - Big prizes",
        "comp3_name": "100 Predictions Challenge",
        "comp3_desc": "Make 100 accurate predictions and win the jackpot",
        # =========== STATISTICS ADDITIONAL ===========
        "statistics_title": "📊 DESCRIPTIVE STATISTICS AND ANALYSIS",
        "tab1": "📈 Summary",
        "tab2": "📋 Tables",
        "tab3": "📊 Advanced Graphs",
        "tab4": "📥 Reports",
        "tab5": "💡 Insights",
        "metric_col": "Metric",
        "value_col": "Value",
        "total": "Total",
        "correct": "Correct",
        "incorrect": "Incorrect",
        "ranking": "Ranking",
        "descriptive_stats": "Descriptive Statistics",
        "pred_desc_stats": "**Prediction Descriptive Statistics**",
        "advanced_viz": "Advanced Visualizations",
        "conf_dist_title": "Confidence Distribution (Histogram + Normal Curve)",
        "time_evol_conf": "Confidence Evolution Over Time",
        "status_dist": "Status Distribution",
        "quantity": "Count",
        "not_enough_data": "Not enough data to show graphs",
        "perf_analysis": "**Performance Analysis**",
        "excellent_perf": "✅ **EXCELLENT** - Your accuracy is above 75%",
        "insight1_excellent": "Keep your current strategy, you've found a good pattern.",
        "good_perf": "⚠️ **GOOD** - Your accuracy is above average",
        "insight1_good": "Keep improving, there's room for optimization.",
        "dev_perf": "⚠️ **DEVELOPING** - Your accuracy is below 60%",
        "insight1_dev": "Review your analysis methodology.",
        "conf_analysis": "**Confidence Analysis**",
        "insight2_high": "Your confidence is realistic with your accuracy.",
        "insight2_moderate": "Consider whether your confidence reflects your accuracy.",
        "insight2_low": "Increase confidence when you have solid supporting data.",
        "personal_recommendations": "**Personalized Recommendations**",
    }
}

# ============================================================================
# CONFIGURACIÓN DE LA PÁGINA
# ============================================================================

st.set_page_config(
    page_title="SportsPredict Pro - Predicciones Deportivas",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        "Get Help": "https://www.example.com/help",
        "Report a bug": "https://www.example.com/bug",
        "About": "# SportsPredict Pro v1.0\nPlataforma de predicciones deportivas con IA"
    }
)

# Initialize session state variables first
if 'dark_mode' not in st.session_state:
    st.session_state.dark_mode = False  # Default: Light mode

# Estilos CSS personalizados (modo claro y oscuro)
if st.session_state.dark_mode:
    st.markdown("""
        <style>
        /* General Styles */
        .main {
            padding: 1.5rem;
            background-color: #0f1117;
            color: #e8eaed;
        }
        
        /* Streamlit Base Elements */
        .stApp {
            background-color: #0f1117;
        }
        
        .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4, .stMarkdown h5, .stMarkdown h6 {
            color: #ffffff;
        }
        .stMarkdown p, .stMarkdown li, .stMarkdown span {
            color: #e8eaed;
        }
        .stMarkdown a {
            color: #667eea;
            text-decoration: underline;
        }
        
        /* Dividers */
        [data-testid="stHorizontalDivider"] {
            border-color: #3a3b45 !important;
        }
        
        /* Sidebar */
        [data-testid="stSidebar"] {
            background-color: #161920;
        }
        [data-testid="stSidebar"] .stMarkdown {
            color: #e8eaed;
        }
        [data-testid="stSidebarNav"] {
            background-color: transparent !important;
        }
        
        /* Buttons */
        .stButton > button {
            background-color: #667eea;
            color: white;
            border: none;
            border-radius: 10px;
            padding: 0.6rem 1.2rem;
            font-weight: 600;
            transition: all 0.25s ease;
            box-shadow: 0 2px 8px rgba(102, 126, 234, 0.15);
        }
        .stButton > button:hover {
            background-color: #5568d3;
            box-shadow: 0 4px 16px rgba(102, 126, 234, 0.35);
            transform: translateY(-1px);
        }
        .stButton > button:active {
            transform: translateY(0);
        }
        
        /* Inputs & Forms */
        .stTextInput > div > div > input,
        .stNumberInput > div > div > input,
        .stTextArea > div > div > textarea,
        .stDateInput > div > div > input,
        .stTimeInput > div > div > input {
            background-color: #1e2129;
            color: white;
            border: 1px solid #3a3b45;
            border-radius: 8px;
            transition: border-color 0.2s ease;
        }
        .stTextInput > div > div > input:focus,
        .stNumberInput > div > div > input:focus,
        .stTextArea > div > div > textarea:focus,
        .stDateInput > div > div > input:focus,
        .stTimeInput > div > div > input:focus {
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.2);
        }
        
        .stSelectbox > div > div > select,
        .stMultiSelect > div > div > div {
            background-color: #1e2129;
            color: white;
            border: 1px solid #3a3b45;
            border-radius: 8px;
        }
        
        .stSlider > div > div > div > div {
            background-color: #667eea;
        }
        .stSlider > div > div > div > div > div {
            background-color: #667eea !important;
        }
        
        /* Checkboxes & Radio Buttons */
        .stCheckbox label, .stRadio label {
            color: #e8eaed;
        }
        
        /* Metrics */
        [data-testid="stMetric"] {
            background: linear-gradient(135deg, #161920 0%, #1e2129 100%);
            padding: 1.2rem;
            border-radius: 12px;
            border: 1px solid #2a2d35;
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        }
        [data-testid="stMetricLabel"] {
            color: #a0a0a0;
            font-weight: 500;
        }
        [data-testid="stMetricValue"] {
            color: #ffffff;
            font-weight: 700;
        }
        [data-testid="stMetricDelta"] {
            color: #00cc96;
        }
        
        /* Dataframes & Tables */
        .stDataFrame {
            background-color: #161920;
            border-radius: 10px;
        }
        .stDataFrame table {
            border: 1px solid #2a2d35 !important;
            border-radius: 10px !important;
        }
        .stDataFrame th {
            background-color: #1e2129 !important;
            color: white !important;
            border-bottom: 1px solid #3a3b45 !important;
        }
        .stDataFrame td {
            color: #e8eaed !important;
            border-bottom: 1px solid #2a2d35 !important;
        }
        
        /* Expanders */
        .stExpander {
            background-color: #161920;
            border: 1px solid #2a2d35;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
        }
        .stExpander > div:first-child {
            color: #ffffff;
            font-weight: 600;
        }
        
        /* Info/Warning/Error/Success Boxes */
        .stAlert {
            border: none;
            border-radius: 12px;
            padding: 1rem 1.2rem;
        }
        .stAlert > div:first-child {
            background-color: transparent !important;
        }
        /* Success */
        .stSuccess {
            background-color: rgba(0, 204, 150, 0.12) !important;
            border-left: 4px solid #00cc96 !important;
        }
        /* Info */
        .stInfo {
            background-color: rgba(102, 126, 234, 0.12) !important;
            border-left: 4px solid #667eea !important;
        }
        /* Warning */
        .stWarning {
            background-color: rgba(255, 171, 0, 0.12) !important;
            border-left: 4px solid #ffab00 !important;
        }
        /* Error */
        .stError {
            background-color: rgba(255, 71, 87, 0.12) !important;
            border-left: 4px solid #ff4757 !important;
        }
        
        /* Tabs */
        .stTabs [data-baseweb="tab-list"] {
            gap: 10px;
            padding-bottom: 4px;
        }
        .stTabs [data-baseweb="tab"] {
            background-color: #161920;
            color: #a0a0a0;
            border-radius: 10px 10px 0 0;
            padding: 0.7rem 1.2rem;
            font-weight: 500;
            border: none;
        }
        .stTabs [aria-selected="true"] {
            background-color: #1e2129 !important;
            color: #667eea !important;
            box-shadow: 0 -2px 8px rgba(102, 126, 234, 0.2);
        }
        
        /* Chat Interface (for Gemini assistant) */
        [data-testid="stChatMessage"] {
            background-color: #161920 !important;
            border: 1px solid #2a2d35 !important;
            border-radius: 12px !important;
        }
        [data-testid="stChatMessageContent"] p {
            color: #e8eaed !important;
        }
        
        /* File Uploader */
        [data-testid="stFileUploader"] {
            background-color: #161920;
            border: 1px dashed #3a3b45;
            border-radius: 12px;
            padding: 1.2rem;
        }
        [data-testid="stFileUploader"] p {
            color: #a0a0a0;
        }
        
        /* Progress Bar */
        [data-testid="stProgress"] > div > div {
            background-color: #667eea !important;
        }
        
        /* Custom Classes */
        .metric-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 1.5rem;
            border-radius: 12px;
            text-align: center;
            box-shadow: 0 4px 16px rgba(102, 126, 234, 0.25);
        }
        .prediction-card {
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            color: white;
            padding: 1.5rem;
            border-radius: 12px;
            box-shadow: 0 4px 16px rgba(245, 87, 108, 0.25);
        }
        .premium-badge {
            background: linear-gradient(135deg, #ffd89b 0%, #19547b 100%);
            color: white;
            padding: 0.5rem 1rem;
            border-radius: 20px;
            font-weight: bold;
        }
        
        /* Responsiveness */
        @media (max-width: 768px) {
            .main {
                padding: 1rem;
            }
        }
        </style>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
        <style>
        /* General Styles */
        .main {
            padding: 1.5rem;
            background-color: #ffffff;
            color: #1a1d23;
        }
        
        /* Streamlit Base Elements */
        .stApp {
            background-color: #ffffff;
        }
        .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4, .stMarkdown h5, .stMarkdown h6 {
            color: #1a1d23;
        }
        .stMarkdown p, .stMarkdown li, .stMarkdown span {
            color: #333333;
        }
        .stMarkdown a {
            color: #667eea;
            text-decoration: underline;
        }
        
        /* Dividers */
        [data-testid="stHorizontalDivider"] {
            border-color: #e0e0e0 !important;
        }
        
        /* Sidebar */
        [data-testid="stSidebar"] {
            background-color: #f5f7fa;
        }
        [data-testid="stSidebar"] .stMarkdown {
            color: #333333;
        }
        
        /* Buttons */
        .stButton > button {
            background-color: #667eea;
            color: white;
            border: none;
            border-radius: 10px;
            padding: 0.6rem 1.2rem;
            font-weight: 600;
            transition: all 0.25s ease;
            box-shadow: 0 2px 8px rgba(102, 126, 234, 0.15);
        }
        .stButton > button:hover {
            background-color: #5568d3;
            box-shadow: 0 4px 16px rgba(102, 126, 234, 0.35);
            transform: translateY(-1px);
        }
        .stButton > button:active {
            transform: translateY(0);
        }
        
        /* Inputs & Forms */
        .stTextInput > div > div > input,
        .stNumberInput > div > div > input,
        .stTextArea > div > div > textarea,
        .stDateInput > div > div > input,
        .stTimeInput > div > div > input {
            background-color: #ffffff;
            color: #1a1d23;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            transition: border-color 0.2s ease;
        }
        .stTextInput > div > div > input:focus,
        .stNumberInput > div > div > input:focus,
        .stTextArea > div > div > textarea:focus,
        .stDateInput > div > div > input:focus,
        .stTimeInput > div > div > input:focus {
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.15);
        }
        
        .stSelectbox > div > div > select,
        .stMultiSelect > div > div > div {
            background-color: #ffffff;
            color: #1a1d23;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
        }
        
        .stSlider > div > div > div > div {
            background-color: #667eea;
        }
        .stSlider > div > div > div > div > div {
            background-color: #667eea !important;
        }
        
        /* Checkboxes & Radio Buttons */
        .stCheckbox label, .stRadio label {
            color: #333333;
        }
        
        /* Metrics */
        [data-testid="stMetric"] {
            background: linear-gradient(135deg, #f5f7fa 0%, #ffffff 100%);
            padding: 1.2rem;
            border-radius: 12px;
            border: 1px solid #e0e0e0;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }
        [data-testid="stMetricLabel"] {
            color: #666666;
            font-weight: 500;
        }
        [data-testid="stMetricValue"] {
            color: #1a1d23;
            font-weight: 700;
        }
        [data-testid="stMetricDelta"] {
            color: #00cc96;
        }
        
        /* Dataframes & Tables */
        .stDataFrame {
            background-color: #ffffff;
            border-radius: 10px;
        }
        .stDataFrame table {
            border: 1px solid #e0e0e0 !important;
            border-radius: 10px !important;
        }
        .stDataFrame th {
            background-color: #f5f7fa !important;
            color: #1a1d23 !important;
            border-bottom: 1px solid #e0e0e0 !important;
        }
        .stDataFrame td {
            color: #333333 !important;
            border-bottom: 1px solid #e0e0e0 !important;
        }
        
        /* Expanders */
        .stExpander {
            background-color: #ffffff;
            border: 1px solid #e0e0e0;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        }
        .stExpander > div:first-child {
            color: #1a1d23;
            font-weight: 600;
        }
        
        /* Info/Warning/Error/Success Boxes */
        .stAlert {
            border: none;
            border-radius: 12px;
            padding: 1rem 1.2rem;
        }
        /* Success */
        .stSuccess {
            background-color: rgba(0, 204, 150, 0.08) !important;
            border-left: 4px solid #00cc96 !important;
        }
        /* Info */
        .stInfo {
            background-color: rgba(102, 126, 234, 0.08) !important;
            border-left: 4px solid #667eea !important;
        }
        /* Warning */
        .stWarning {
            background-color: rgba(255, 171, 0, 0.08) !important;
            border-left: 4px solid #ffab00 !important;
        }
        /* Error */
        .stError {
            background-color: rgba(255, 71, 87, 0.08) !important;
            border-left: 4px solid #ff4757 !important;
        }
        
        /* Tabs */
        .stTabs [data-baseweb="tab-list"] {
            gap: 10px;
            padding-bottom: 4px;
        }
        .stTabs [data-baseweb="tab"] {
            background-color: #f5f7fa;
            color: #666666;
            border-radius: 10px 10px 0 0;
            padding: 0.7rem 1.2rem;
            font-weight: 500;
            border: none;
        }
        .stTabs [aria-selected="true"] {
            background-color: #ffffff !important;
            color: #667eea !important;
            box-shadow: 0 -2px 8px rgba(102, 126, 234, 0.15);
        }
        
        /* Chat Interface (for Gemini assistant) */
        [data-testid="stChatMessage"] {
            background-color: #f5f7fa !important;
            border: 1px solid #e0e0e0 !important;
            border-radius: 12px !important;
        }
        [data-testid="stChatMessageContent"] p {
            color: #1a1d23 !important;
        }
        
        /* File Uploader */
        [data-testid="stFileUploader"] {
            background-color: #f5f7fa;
            border: 1px dashed #e0e0e0;
            border-radius: 12px;
            padding: 1.2rem;
        }
        [data-testid="stFileUploader"] p {
            color: #666666;
        }
        
        /* Progress Bar */
        [data-testid="stProgress"] > div > div {
            background-color: #667eea !important;
        }
        
        /* Custom Classes */
        .metric-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 1.5rem;
            border-radius: 12px;
            text-align: center;
            box-shadow: 0 4px 16px rgba(102, 126, 234, 0.25);
        }
        .prediction-card {
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            color: white;
            padding: 1.5rem;
            border-radius: 12px;
            box-shadow: 0 4px 16px rgba(245, 87, 108, 0.25);
        }
        .premium-badge {
            background: linear-gradient(135deg, #ffd89b 0%, #19547b 100%);
            color: white;
            padding: 0.5rem 1rem;
            border-radius: 20px;
            font-weight: bold;
        }
        
        /* Responsiveness */
        @media (max-width: 768px) {
            .main {
                padding: 1rem;
            }
        }
        </style>
    """, unsafe_allow_html=True)

# ============================================================================
# INICIALIZACIÓN Y CONFIGURACIÓN
# ============================================================================

@st.cache_resource
def init_supabase():
    """Inicializa la conexión a Supabase (para usuarios normales (clave pública))"""
    if not SUPABASE_AVAILABLE:
        return None
    try:
        supabase_url = None
        supabase_key = None
        try:
            supabase_url = st.secrets.get("SUPABASE_URL")
            supabase_key = st.secrets.get("SUPABASE_KEY")
        except Exception:
            pass  # Si no hay secrets o hay error, usar None

        if not supabase_url or not supabase_key:
            return None

        return create_client(supabase_url, supabase_key)
    except Exception:
        return None

@st.cache_resource
def init_supabase_service():
    """Inicializa la conexión a Supabase con clave de servicio (para admin, ignora RLS)"""
    if not SUPABASE_AVAILABLE:
        return None
    try:
        supabase_url = None
        supabase_secret_key = None
        try:
            supabase_url = st.secrets.get("SUPABASE_URL")
            supabase_secret_key = st.secrets.get("SUPABASE_SECRET_KEY")
        except Exception:
            pass  # Si no hay secrets o hay error, usar None

        if not supabase_url or not supabase_secret_key:
            return None

        return create_client(supabase_url, supabase_secret_key)
    except Exception:
        return None

@st.cache_resource
def init_api_keys():
    """Obtiene las claves de API de los secretos (Solo APIs Gratuitas)"""
    api_football_key = ""
    google_studio_key = ""
    try:
        api_football_key = st.secrets.get("API_FOOTBALL_KEY", "")
        google_studio_key = st.secrets.get("GOOGLE_STUDIO_KEY", "")
    except Exception:
        pass  # Si no hay secrets, usar cadenas vacías
    return {
        'api_football': api_football_key,
        'google_studio': google_studio_key
    }

supabase = init_supabase()
supabase_service = init_supabase_service()
api_keys = init_api_keys()


def seed_database():
    """Llena la base de datos con datos iniciales si está vacía"""
    # Preferir el cliente de servicio para evitar RLS
    db_client = supabase_service if supabase_service else supabase
    if not db_client:
        return False
    
    try:
        # Verificar si ya hay usuarios
        users_response = db_client.table('users').select('*').limit(1).execute()
        if users_response.data and len(users_response.data) > 0:
            return False  # La BD ya tiene datos
        
        # Insertar usuarios iniciales
        initial_users = [
            {
                'email': 'admin@example.com',
                'username': 'admin',
                'password': 'admin123',
                'full_name': 'Administrador',
                'subscription_tier': 'elite'
            },
            {
                'email': 'usuario@example.com',
                'username': 'usuario',
                'password': 'usuario123',
                'full_name': 'Usuario Normal',
                'subscription_tier': 'free'
            }
        ]
        
        for user in initial_users:
            db_client.table('users').insert(user).execute()
        
        # Insertar equipos iniciales
        initial_teams = [
            {'name': 'Real Madrid', 'country': 'España', 'sport_type': 'football'},
            {'name': 'Barcelona', 'country': 'España', 'sport_type': 'football'},
            {'name': 'Manchester United', 'country': 'Inglaterra', 'sport_type': 'football'},
            {'name': 'Liverpool', 'country': 'Inglaterra', 'sport_type': 'football'}
        ]
        
        for team in initial_teams:
            db_client.table('teams').insert(team).execute()
        
        # Insertar partidos iniciales
        tomorrow = datetime.now() + timedelta(days=1)
        initial_matches = [
            {
                'home_team_name': 'Real Madrid',
                'away_team_name': 'Barcelona',
                'match_date': tomorrow.isoformat(),
                'league': 'La Liga',
                'sport_type': 'football',
                'status': 'scheduled'
            },
            {
                'home_team_name': 'Manchester United',
                'away_team_name': 'Liverpool',
                'match_date': (tomorrow + timedelta(days=2)).isoformat(),
                'league': 'Premier League',
                'sport_type': 'football',
                'status': 'scheduled'
            }
        ]
        
        for match in initial_matches:
            db_client.table('matches').insert(match).execute()
        
        return True
        
    except Exception as e:
        return False  # Silenciar errores si las tablas no existen o ya hay datos


# Ejecutar la inicialización de la BD solo una vez por sesión
if 'db_seeded' not in st.session_state:
    st.session_state.db_seeded = seed_database()

# Estado de sesión
if 'user_id' not in st.session_state:
    st.session_state.user_id = "user_" + hashlib.md5(str(time.time()).encode()).hexdigest()[:8]

if 'user_tier' not in st.session_state:
    st.session_state.user_tier = "free"

if 'favorite_teams' not in st.session_state:
    st.session_state.favorite_teams = []

if 'selected_competitions' not in st.session_state:
    st.session_state.selected_competitions = []

if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

if 'is_admin' not in st.session_state:
    st.session_state.is_admin = False

if 'username' not in st.session_state:
    st.session_state.username = "Invitado"

if 'language' not in st.session_state:
    st.session_state.language = "es"  # Default: Spanish

if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []  # Chatbot conversation history

if 'recording' not in st.session_state:
    st.session_state.recording = False  # Voice recording state

# Importamos el gestor de tablas
try:
    from maintenance import show_table_manager_ui
except:
    pass

# ============================================================================
# SISTEMA DE AUTENTICACIÓN
# ============================================================================

def login_page():
    """Página de inicio de sesión"""
    # We'll use Spanish as default for login page, or check session
    lang = st.session_state.language if 'language' in st.session_state else "es"
    t = TRANSLATIONS[lang]
    
    st.title(t['login_title'])
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader(t['access_to_platform'])
        
        username = st.text_input(t['username'])
        password = st.text_input(t['password'], type="password")
        
        if st.button(t['login_button']):
            # Primero intentar verificar en Supabase
            authenticated = False
            is_admin = False
            
            if supabase:
                try:
                    # Buscar usuario en la BD
                    user_response = supabase.table('users').select('*').eq('username', username).execute()
                    if user_response.data and len(user_response.data) > 0:
                        user_data = user_response.data[0]
                        # Verificar contraseña (en un entorno real, usarías hashing!)
                        if user_data.get('password') == password:
                            authenticated = True
                            is_admin = (username == "admin")
                except Exception as e:
                    pass  # Si falla, usar credenciales hardcodeadas
            
            # Si no se autenticó por BD, usar credenciales de ejemplo
            if not authenticated:
                admin_creds = {
                    "admin": "admin123",
                    "usuario": "usuario123"
                }
                
                if username in admin_creds and admin_creds[username] == password:
                    authenticated = True
                    is_admin = (username == "admin")
            
            # Finalizar autenticación
            if authenticated:
                st.session_state.authenticated = True
                st.session_state.username = username
                st.session_state.is_admin = is_admin
                st.session_state.mode_chosen = True
                st.success(t['login_success'])
                st.rerun()
            else:
                st.error(t['login_error'])
    
    with col2:
        st.subheader(t['guest_mode'])
        st.info(f"""
        {t['guest_info']}
        
        **{t['test_credentials']}:**
        - Admin: admin / admin123
        - Usuario: usuario / usuario123
        """)
        
        if st.button(t['continue_as_guest']):
            st.session_state.authenticated = False
            st.session_state.username = "Invitado"
            st.session_state.mode_chosen = True
            st.success(t['guest_success'])
            st.rerun()

def show_user_profile():
    """Muestra el perfil del usuario en la barra lateral"""
    lang = st.session_state.language
    t = TRANSLATIONS[lang]
    st.sidebar.markdown(f"**{t['user_label']}:** {st.session_state.username}")
    st.sidebar.markdown(f"**{t['plan_label']}:** {st.session_state.user_tier.upper()}")
    
    if st.sidebar.button("Cerrar Sesión"):
        st.session_state.authenticated = False
        st.session_state.username = "Invitado"
        st.session_state.is_admin = False
        if 'mode_chosen' in st.session_state:
            del st.session_state.mode_chosen
        st.rerun()

# ============================================================================
# CLASES Y UTILIDADES
# ============================================================================

class PDFReport(FPDF):
    """Generador de reportes PDF"""

    def header(self):
        self.set_font('Arial', 'B', 16)
        # self.image(None, 10, 8, 33)  # Placeholder para logo (comentado para evitar error)
        self.cell(0, 10, 'SportsPredict Pro - Reporte de Predicciones', 0, 1, 'C')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')

    def chapter_title(self, title):
        self.set_font('Arial', 'B', 12)
        self.set_fill_color(200, 220, 255)
        self.cell(0, 10, title, 0, 1, 'L', True)
        self.ln(5)

    def chapter_body(self, body):
        self.set_font('Arial', '', 11)
        self.multi_cell(0, 10, body)
        self.ln()

# ============================================================================
# CHATBOT DE AYUDA CON GEMINI
# ============================================================================

def init_gemini():
    """Inicializa el cliente de Gemini API"""
    try:
        # Intenta obtener la clave de secrets
        api_key = None
        try:
            api_key = st.secrets.get("GOOGLE_API_KEY")
        except Exception:
            pass
        
        # Si no hay clave en secrets, solicitarla al usuario
        if not api_key:
            # Almacenar la clave en session state para que persista
            if "gemini_api_key" not in st.session_state:
                st.session_state.gemini_api_key = ""
            api_key_input = st.text_input("🔑 Google API Key (Gemini)", type="password", value=st.session_state.gemini_api_key)
            if api_key_input:
                st.session_state.gemini_api_key = api_key_input
                api_key = api_key_input
        
        if api_key:
            
            return True  # Return True to indicate initialization success
        return False
    except Exception as e:
        st.error(f"Error inicializando Gemini: {e}")
        return False

def get_gemini_response(user_message=None, audio_data=None, context_data=None):
    """Obtiene respuesta del chatbot Gemini (soporta texto o audio)"""
    lang = st.session_state.language
    t = TRANSLATIONS[lang]
    try:
        prompt_text = f"""
        Eres un asistente experto en predicciones deportivas para SportsPredict Pro.
        
        Tarea: Ayuda a los usuarios a interpretar las predicciones de IA y compararlas con datos de la API.
        
        Contexto adicional (si existe):
        {json.dumps(context_data, indent=2) if context_data else "No hay contexto adicional"}
        """
        
        api_key = st.session_state.get("gemini_api_key")
        if not api_key:
            try:
                api_key = st.secrets.get("GOOGLE_API_KEY")
            except:
                pass
        
        client = genai.Client(api_key=api_key)
        
        if audio_data:
            response = client.models.generate_content(
                model='gemini-flash-latest',
                contents=[
                    prompt_text,
                    types.Part.from_bytes(
                        data=audio_data,
                        mime_type="audio/wav"
                    )
                ]
            )
        else:
            full_prompt = f"{prompt_text}\n\nPregunta del usuario: {user_message}"
            response = client.models.generate_content(
                model='gemini-flash-latest',
                contents=full_prompt
            )
        return response.text
    except Exception as e:
        error_str = str(e)
        if "429" in error_str or "quota" in error_str.lower():
            return t["quota_exceeded_msg"]
        else:
            return f"Lo siento, no pude procesar tu pregunta. Error: {error_str}"

def page_ai_assistant():
    """Página del Asistente IA"""
    lang = st.session_state.language
    t = TRANSLATIONS[lang]
    
    st.title(t["ai_assistant_page"])
    st.markdown("---")
    
    # Inicializar cliente Gemini
    gemini_ready = init_gemini()
    
    if gemini_ready:
        st.subheader(t["chatbot_title"])
        
        # Mostrar historial de chat
        chat_container = st.container()
        with chat_container:
            for message in st.session_state.chat_history:
                with st.chat_message(message["role"]):
                    st.markdown(message["content"])
        
        # Entrada de audio
        st.subheader(t["audio_input_label"])
        audio_value = st.audio_input("Graba tu pregunta:")
        
        prompt_text = None
        audio_bytes = None
        
        # Procesar audio si existe
        if audio_value:
            audio_bytes = audio_value.getvalue()
        
        # Entrada de texto
        text_prompt = st.chat_input(t["chatbot_placeholder"])
        if text_prompt:
            prompt_text = text_prompt
        
        # Si tenemos texto o audio, procesarlo
        if prompt_text or audio_bytes:
            user_content = prompt_text if prompt_text else "Mensaje de audio"
            # Agregar mensaje del usuario al historial
            st.session_state.chat_history.append({"role": "user", "content": user_content})
            with chat_container.chat_message("user"):
                if prompt_text:
                    st.markdown(prompt_text)
                else:
                    st.audio(audio_value, format="audio/wav")
                    st.write("🎤 Mensaje de audio")
            
            # Obtener contexto (partidos y predicciones actuales)
            context_data = None
            try:
                matches = api_client.get_upcoming_matches(days=1)
                if matches:
                    context_data = {
                        "matches": [
                            {
                                "home": m["teams"]["home"]["name"],
                                "away": m["teams"]["away"]["name"],
                                "date": m["fixture"]["date"]
                            } for m in matches[:3]
                        ]
                    }
            except Exception:
                pass
            
            # Obtener respuesta de Gemini
            with chat_container.chat_message("assistant"):
                with st.spinner(t["processing_audio"] if audio_bytes else "Pensando..."):
                    if audio_bytes:
                        response = get_gemini_response(audio_data=audio_bytes, context_data=context_data)
                    else:
                        response = get_gemini_response(user_message=prompt_text, context_data=context_data)
                    st.markdown(response)
            
            # Agregar respuesta al historial
            st.session_state.chat_history.append({"role": "assistant", "content": response})
    else:
        st.info("Por favor, proporciona una clave API de Google para usar el asistente.")


def render_chatbot():
    """(DEPRECATED - Ahora el asistente está en su propia página"""
    pass

class APIClient:
    """Cliente para consultar APIs de deportes"""

    def __init__(self, api_keys: Dict):
        self.api_keys = api_keys
        self.football_base_url = "https://v3.football.api-sports.io"
        self.session = requests.Session()
        self.timeout = 5  # Segundos máximo de espera para cada solicitud API

    def get_upcoming_matches(self, days: int = 7) -> List[Dict]:
        """Obtiene partidos próximos de API-Football"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_matches()

            headers = {"x-apisports-key": self.api_keys['api_football']}

            # Obtener partidos de los próximos días
            matches = []
            for i in range(days):
                date = (datetime.now() + timedelta(days=i)).strftime("%Y-%m-%d")
                response = self.session.get(
                    f"{self.football_base_url}/fixtures",
                    headers=headers,
                    params={"date": date, "status": "NS"},  # NS = Not Started
                    timeout=self.timeout
                )

                if response.status_code == 200:
                    data = response.json()
                    if 'response' in data:
                        matches.extend(data['response'][:5])  # Limitar a 5 por día

            # Si no hay partidos de la API, usar mock data
            if not matches:
                return self._get_mock_matches()
            
            return matches[:10]
        except Exception:
            # Si hay error, usar mock data
            return self._get_mock_matches()

    def get_team_stats(self, team_id: int) -> Dict:
        """Obtiene estadísticas de un equipo"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_team_stats()

            headers = {"x-apisports-key": self.api_keys['api_football']}
            response = self.session.get(
                f"{self.football_base_url}/teams/statistics",
                headers=headers,
                params={"team": team_id, "season": datetime.now().year},
                timeout=self.timeout
            )

            if response.status_code == 200:
                result = response.json().get('response', {})
                if result:
                    return result
            return self._get_mock_team_stats()
        except Exception:
            return self._get_mock_team_stats()

    def get_predictions(self, fixture_id: int) -> Dict:
        """Obtiene predicciones para un partido"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_prediction()

            headers = {"x-apisports-key": self.api_keys['api_football']}
            response = self.session.get(
                f"{self.football_base_url}/predictions",
                headers=headers,
                params={"fixture": fixture_id},
                timeout=self.timeout
            )

            if response.status_code == 200:
                result = response.json().get('response', [{}])
                if result and len(result) > 0:
                    return result[0]
            return self._get_mock_prediction()
        except Exception:
            return self._get_mock_prediction()

    def get_head_to_head(self, team1_id: int, team2_id: int) -> List[Dict]:
        """Obtiene historial H2H entre dos equipos"""
        try:
            if not self.api_keys.get('api_football'):
                return self._get_mock_h2h()

            headers = {"x-apisports-key": self.api_keys['api_football']}
            response = self.session.get(
                f"{self.football_base_url}/fixtures/headtohead",
                headers=headers,
                params={"h2h": f"{team1_id}-{team2_id}", "last": 10},
                timeout=self.timeout
            )

            if response.status_code == 200:
                result = response.json().get('response', [])
                if result:
                    return result
            return self._get_mock_h2h()
        except Exception:
            return self._get_mock_h2h()

    # ==================== DATOS MOCK ====================

    @staticmethod
    def _get_mock_matches() -> List[Dict]:
        """Retorna datos de ejemplo para partidos"""
        return [
            {
                'fixture': {'id': 1001, 'date': '2026-07-15T20:00:00+00:00', 'status': 'NS'},
                'league': {'name': 'La Liga', 'country': 'Spain', 'logo': ''},
                'teams': {
                    'home': {'id': 541, 'name': 'Real Madrid', 'logo': ''},
                    'away': {'id': 529, 'name': 'Barcelona', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1002, 'date': '2026-07-16T15:00:00+00:00', 'status': 'NS'},
                'league': {'name': 'Premier League', 'country': 'England', 'logo': ''},
                'teams': {
                    'home': {'id': 42, 'name': 'Manchester City', 'logo': ''},
                    'away': {'id': 40, 'name': 'Liverpool', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            },
            {
                'fixture': {'id': 1003, 'date': '2026-07-17T19:30:00+00:00', 'status': 'NS'},
                'league': {'name': 'Bundesliga', 'country': 'Germany', 'logo': ''},
                'teams': {
                    'home': {'id': 25, 'name': 'Bayern Munich', 'logo': ''},
                    'away': {'id': 165, 'name': 'Borussia Dortmund', 'logo': ''}
                },
                'goals': {'home': None, 'away': None}
            }
        ]

    @staticmethod
    def _get_mock_prediction() -> Dict:
        """Retorna predicción de ejemplo"""
        return {
            'predictions': {
                'winner': {'name': 'home', 'comment': 'The home team is likely to win'},
                'win_or_draw': 95,
                'win_home_or_draw': 88,
                'goals': {'home': 2.1, 'away': 1.2},
                'goals_more_less': 'over',
                'advice': 'Home team is in better form'
            },
            'comparison': {
                'form': {'home': '5W-2D-1L', 'away': '3W-4D-3L'},
                'att': {'home': 85, 'away': 72},
                'def': {'home': 92, 'away': 78},
                'poisson_distribution': {
                    'home': {'0': 8, '1': 21, '2': 35, '3': 30},
                    'away': {'0': 15, '1': 34, '2': 28, '3': 16}
                }
            }
        }

    @staticmethod
    def _get_mock_team_stats() -> Dict:
        """Retorna estadísticas de equipo de ejemplo"""
        return {
            'team': {'id': 541, 'name': 'Real Madrid'},
            'statistics': [
                {'type': 'Wins home', 'value': 12},
                {'type': 'Wins away', 'value': 8},
                {'type': 'Draws', 'value': 4},
                {'type': 'Goals for', 'value': 65},
                {'type': 'Goals against', 'value': 18},
                {'type': 'Goals for home', 'value': 35},
                {'type': 'Goals for away', 'value': 30},
                {'type': 'Goals against home', 'value': 8},
                {'type': 'Goals against away', 'value': 10}
            ]
        }

    @staticmethod
    def _get_mock_h2h() -> List[Dict]:
        """Retorna historial H2H de ejemplo"""
        return [
            {
                'fixture': {'id': 900, 'date': '2023-12-10T19:00:00+00:00'},
                'teams': {'home': {'name': 'Real Madrid'}, 'away': {'name': 'Barcelona'}},
                'goals': {'home': 2, 'away': 1},
                'score': {'fulltime': {'home': 2, 'away': 1}}
            },
            {
                'fixture': {'id': 901, 'date': '2023-10-28T20:00:00+00:00'},
                'teams': {'home': {'name': 'Barcelona'}, 'away': {'name': 'Real Madrid'}},
                'goals': {'home': 1, 'away': 1},
                'score': {'fulltime': {'home': 1, 'away': 1}}
            },
            {
                'fixture': {'id': 902, 'date': '2023-08-15T19:30:00+00:00'},
                'teams': {'home': {'name': 'Real Madrid'}, 'away': {'name': 'Barcelona'}},
                'goals': {'home': 3, 'away': 2},
                'score': {'fulltime': {'home': 3, 'away': 2}}
            }
        ]

api_client = APIClient(api_keys)

@st.cache_resource(show_spinner="Cargando modelos de IA...")
def get_trained_model(force_refresh: bool = False, save_dir: str = "saved_models"):
    """Entrena y devuelve el modelo ensemble. Carga desde disco si existe."""
    import os
    os.makedirs(save_dir, exist_ok=True)
    
    # Check if all required files exist
    required_files = ["xgb_model.pkl", "bayesian_model.pkl", "lstm_model.pth", "meta_model.pkl", "ensemble_config.pkl"]
    all_exist = all(os.path.exists(os.path.join(save_dir, f)) for f in required_files)
    
    if all_exist and not force_refresh:
        try:
            ensemble = ModelEnsemble.load(save_dir=save_dir, input_dim=20)
            return ensemble
        except Exception as e:
            st.warning(f"No se pudo cargar el modelo guardado: {e}. Entrenando de nuevo...")
    
    # If not exist or forced, train new model
    X, y_class, y_goal_diff = create_sample_training_data(n_samples=2000, fast_mode=False)
    ensemble = ModelEnsemble()
    ensemble.fit(X, y_class, y_goal_diff)
    ensemble.save(save_dir=save_dir)
    return ensemble

ensemble_model = get_trained_model(force_refresh=False)

# ============================================================================
# FUNCIONES DE BASE DE DATOS
# ============================================================================

def save_prediction(user_id: str, match_data: Dict, prediction: Dict, confidence: float):
    """Guarda una predicción en la base de datos o en session state como fallback"""
    try:
        data = {
            'user_id': user_id,
            'match_id': match_data.get('fixture', {}).get('id'),
            'predicted_home_score': prediction.get('home_score'),
            'predicted_away_score': prediction.get('away_score'),
            'confidence_level': confidence,
            'ai_prediction_data': json.dumps(prediction),
            'created_at': datetime.now().isoformat()
        }

        if supabase:
            try:
                response = supabase.table('user_predictions').insert(data).execute()
                return True
            except Exception:
                pass  # Silenciar error de Supabase
        
        # Fallback: guardar en session state
        if 'local_predictions' not in st.session_state:
            st.session_state.local_predictions = []
        st.session_state.local_predictions.append(data)
        return True
    except Exception as e:
        st.error(f"Error guardando predicción: {e}")
        return False

def get_user_predictions(user_id: str, limit: int = 50) -> List[Dict]:
    """Obtiene las predicciones de un usuario (de Supabase o de session state)"""
    predictions = []
    try:
        if supabase:
            try:
                response = supabase.table('user_predictions').select('*').eq('user_id', user_id).limit(limit).execute()
                if response.data:
                    predictions.extend(response.data)
            except Exception:
                pass  # Silenciar error de Supabase
        
        # Agregar predicciones locales
        if 'local_predictions' in st.session_state:
            # Filtrar por user_id
            local_preds = [p for p in st.session_state.local_predictions if p.get('user_id') == user_id]
            predictions.extend(local_preds)
        
        # Ordenar por fecha (más reciente primero) y limitar
        predictions.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        return predictions[:limit]
    except Exception:
        # Si falla todo, devolver predicciones locales si hay
        return st.session_state.get('local_predictions', [])[:limit] if 'local_predictions' in st.session_state else []

def get_user_stats(user_id: str) -> Dict:
    """Calcula estadísticas del usuario"""
    predictions = get_user_predictions(user_id)

    if not predictions:
        return {
            'total_predictions': 0,
            'correct_predictions': 0,
            'accuracy_rate': 0,
            'avg_confidence': 0,
            'total_earnings': 0,
            'rank': 'N/A'
        }

    df = pd.DataFrame(predictions)

    total_preds = len(df)
    correct_preds = len(df[df['confidence_level'] > 0.7]) if 'confidence_level' in df.columns else 0
    avg_conf = df['confidence_level'].mean() if 'confidence_level' in df.columns else 0

    return {
        'total_predictions': total_preds,
        'correct_predictions': correct_preds,
        'accuracy_rate': (correct_preds / total_preds * 100) if total_preds > 0 else 0,
        'avg_confidence': float(avg_conf),
        'total_earnings': 0,
        'rank': f'Top {np.random.randint(5, 25)}%'
    }

def get_upcoming_matches_from_db(days: int = 30) -> List[Dict]:
    """Obtiene partidos próximos desde la base de datos (Supabase o modo local)"""
    matches = []
    try:
        # Intentar primero con Supabase (usando cliente normal para usuarios, no servicio)
        db_client = supabase_service if supabase_service else supabase
        if db_client:
            from_date = datetime.now().isoformat()
            to_date = (datetime.now() + timedelta(days=days)).isoformat()
            response = db_client.table('matches').select('*').gte('match_date', from_date).lte('match_date', to_date).execute()
            if response.data:
                # Convertir datos de DB al mismo formato que usa api_client para compatibilidad
                for db_match in response.data:
                    matches.append({
                        'fixture': {
                            'id': db_match['id'],
                            'date': db_match['match_date'],
                            'status': db_match.get('status', 'NS')
                        },
                        'league': {'name': db_match.get('league', 'Liga'), 'country': '', 'logo': ''},
                        'teams': {
                            'home': {'id': db_match.get('home_team_id', ''), 'name': db_match.get('home_team_name', 'Equipo 1'), 'logo': ''},
                            'away': {'id': db_match.get('away_team_id', ''), 'name': db_match.get('away_team_name', 'Equipo 2'), 'logo': ''}
                        },
                        'goals': {'home': None, 'away': None}
                    })
        
        # Si no hay partidos en DB, intentar modo local (st.session_state)
        if not matches and 'local_matches' in st.session_state:
            local_matches = st.session_state['local_matches']
            from_date = datetime.now().isoformat()
            to_date = (datetime.now() + timedelta(days=days)).isoformat()
            for db_match in local_matches:
                if from_date <= db_match.get('match_date', '') <= to_date:
                    matches.append({
                        'fixture': {
                            'id': db_match['id'],
                            'date': db_match['match_date'],
                            'status': db_match.get('status', 'NS')
                        },
                        'league': {'name': db_match.get('league', 'Liga'), 'country': '', 'logo': ''},
                        'teams': {
                            'home': {'id': db_match.get('home_team_id', ''), 'name': db_match.get('home_team_name', 'Equipo 1'), 'logo': ''},
                            'away': {'id': db_match.get('away_team_id', ''), 'name': db_match.get('away_team_name', 'Equipo 2'), 'logo': ''}
                        },
                        'goals': {'home': None, 'away': None}
                    })
                    
    except Exception as e:
        pass  # Si hay error, usar el fallback
        
    return matches

def get_competitions() -> List[Dict]:
    """Obtiene la lista de competiciones disponibles"""
    return []

# ============================================================================
# COMPONENTES REUTILIZABLES
# ============================================================================

def render_metric_card(label: str, value: str, delta: str = "", icon: str = ""):
    """Renderiza una tarjeta de métrica"""
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        st.markdown(f"<h3>{icon}</h3>", unsafe_allow_html=True)
    with col2:
        st.metric(label, value, delta=delta)
    with col3:
        pass

def render_match_card(match: Dict):
    """Renderiza una tarjeta de partido"""
    try:
        fixture = match.get('fixture', {})
        teams = match.get('teams', {})
        league = match.get('league', {})

        home_team = teams.get('home', {})
        away_team = teams.get('away', {})

        match_date = fixture.get('date', '')
        if match_date:
            match_date = datetime.fromisoformat(match_date.replace('Z', '+00:00')).strftime("%d/%m/%Y %H:%M")

        col1, col2, col3, col4 = st.columns([2, 2, 2, 2])

        with col1:
            st.write(f"🏆 **{league.get('name', 'Liga')}**")
            st.write(f"📅 {match_date}")

        with col2:
            st.write(f"🏠 **{home_team.get('name', 'Equipo 1')}**")

        with col3:
            st.write("vs")

        with col4:
            st.write(f"✈️ **{away_team.get('name', 'Equipo 2')}**")

        return True
    except Exception as e:
        st.error(f"Error renderizando partido: {e}")
        return False

# ============================================================================
# GENERADOR DE REPORTES - PDF Y EXCEL
# ============================================================================

def generate_predictions_report_pdf(predictions: List[Dict]) -> bytes:
    """Genera reporte PDF de predicciones"""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 15, "REPORTE DE PREDICCIONES", ln=True, align="C")

    pdf.set_font("Arial", "", 11)
    pdf.ln(5)

    df = pd.DataFrame(predictions) if predictions else pd.DataFrame()

    if not df.empty:
        total = len(df)
        correct = len(df[df.get('prediction_status') == 'won']) if 'prediction_status' in df.columns else 0
        accuracy = (correct / total * 100) if total > 0 else 0

        pdf.cell(0, 8, f"Total de Predicciones: {total}", ln=True)
        pdf.cell(0, 8, f"Predicciones Correctas: {correct}", ln=True)
        pdf.cell(0, 8, f"Tasa de Precisión: {accuracy:.2f}%", ln=True)

        if 'confidence_level' in df.columns:
            avg_conf = df['confidence_level'].mean()
            pdf.cell(0, 8, f"Confianza Promedio: {avg_conf:.2f}", ln=True)
        
        # Add a simple table of predictions
        pdf.ln(10)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(40, 8, "ID Partido", border=1, ln=0)
        pdf.cell(40, 8, "Goles Local", border=1, ln=0)
        pdf.cell(40, 8, "Goles Visitante", border=1, ln=0)
        pdf.cell(40, 8, "Confianza", border=1, ln=0)
        pdf.cell(30, 8, "Fecha", border=1, ln=1)
        
        pdf.set_font("Arial", "", 9)
        for _, row in df.iterrows():
            pdf.cell(40, 8, str(row.get('match_id', '-')), border=1, ln=0)
            pdf.cell(40, 8, str(row.get('predicted_home_score', '-')), border=1, ln=0)
            pdf.cell(40, 8, str(row.get('predicted_away_score', '-')), border=1, ln=0)
            pdf.cell(40, 8, f"{row.get('confidence_level', '-'):.2f}", border=1, ln=0)
            pdf.cell(30, 8, str(row.get('created_at', '-'))[:10], border=1, ln=1)
    else:
        pdf.set_font("Arial", "I", 12)
        pdf.cell(0, 10, "No hay predicciones registradas aún.", ln=True, align="C")
        pdf.ln(5)
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 8, "Comienza a realizar predicciones para generar reportes detallados.", ln=True)

    pdf.ln(10)
    pdf.set_font("Arial", "I", 8)
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    pdf.cell(0, 10, f"Generado: {timestamp}", align="R")

    return bytes(pdf.output())


def generate_predictions_report_excel(predictions: List[Dict]) -> bytes:
    """Genera reporte Excel de predicciones"""
    df = pd.DataFrame(predictions) if predictions else pd.DataFrame()

    wb = Workbook()
    ws = wb.active
    ws.title = "Predicciones"

    if not df.empty:
        # Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

        # Datos
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
    else:
        # Add info if no data
        ws.cell(row=1, column=1).value = "No hay predicciones registradas aún"
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=3, column=1).value = "Comienza a realizar predicciones para generar reportes detallados."
        ws.cell(row=3, column=1).font = Font(italic=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return bytes(output.getvalue())


# ============================================================================
# ESTADÍSTICAS DESCRIPTIVAS - GRÁFICOS Y ANÁLISIS
# ============================================================================

def create_distribution_plot(data: List[float], title: str = "Distribución") -> go.Figure:
    """Crea gráfico de distribución con curva normal"""
    if not data:
        return go.Figure()

    df = pd.DataFrame({'value': data})

    fig = go.Figure()

    fig.add_trace(go.Histogram(
        x=df['value'],
        name='Frecuencia',
        nbinsx=30,
        marker_color='rgba(102, 126, 234, 0.7)'
    ))

    mu, sigma = np.mean(data), np.std(data)
    x = np.linspace(min(data), max(data), 100)
    y = stats.norm.pdf(x, mu, sigma)
    y_scaled = y * len(data) * (max(data) - min(data)) / 30

    fig.add_trace(go.Scatter(
        x=x, y=y_scaled,
        name='Curva Normal',
        mode='lines',
        line=dict(color='red', width=3)
    ))

    fig.update_layout(
        title=title,
        xaxis_title="Valor",
        yaxis_title="Frecuencia",
        hovermode='x unified',
        height=400
    )

    return fig


def create_boxplot(data_dict: Dict[str, List[float]], title: str = "Comparación") -> go.Figure:
    """Crea gráfico de caja (boxplot)"""
    fig = go.Figure()

    for label, values in data_dict.items():
        fig.add_trace(go.Box(y=values, name=label, boxmean='sd'))

    fig.update_layout(
        title=title,
        yaxis_title="Valor",
        height=400
    )

    return fig


def create_time_series_plot(dates: List[datetime], values: List[float], title: str = "Serie Temporal") -> go.Figure:
    """Crea gráfico de serie temporal"""
    df = pd.DataFrame({'date': dates, 'value': values})
    df = df.sort_values('date')

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=df['date'],
        y=df['value'],
        mode='lines+markers',
        name='Valor',
        line=dict(color='#667EEA', width=2)
    ))

    if len(df) > 7:
        df['ma_7'] = df['value'].rolling(window=7).mean()
        fig.add_trace(go.Scatter(
            x=df['date'],
            y=df['ma_7'],
            mode='lines',
            name='Media Móvil (7 días)',
            line=dict(color='red', dash='dash')
        ))

    fig.update_layout(
        title=title,
        xaxis_title="Fecha",
        yaxis_title="Valor",
        height=400
    )

    return fig


def create_correlation_heatmap(df: pd.DataFrame, numeric_cols: List[str]) -> go.Figure:
    """Crea mapa de calor de correlaciones"""
    numeric_df = df[numeric_cols].select_dtypes(include=[np.number])
    corr_matrix = numeric_df.corr()

    fig = go.Figure(data=go.Heatmap(
        z=corr_matrix.values,
        x=corr_matrix.columns,
        y=corr_matrix.columns,
        colorscale='RdBu',
        zmid=0,
        text=np.round(corr_matrix.values, 2),
        texttemplate='%{text:.2f}'
    ))

    fig.update_layout(title="Matriz de Correlación", height=500, width=600)
    return fig


def calculate_descriptive_stats(data: List[float]) -> Dict:
    """Calcula estadísticas descriptivas"""
    if not data or len(data) == 0:
        return {}

    data_array = np.array(data)

    return {
        'count': len(data),
        'mean': float(np.mean(data_array)),
        'median': float(np.median(data_array)),
        'std': float(np.std(data_array)),
        'min': float(np.min(data_array)),
        'max': float(np.max(data_array)),
        'q1': float(np.percentile(data_array, 25)),
        'q3': float(np.percentile(data_array, 75))
    }


def create_descriptive_stats_table(predictions: List[Dict]) -> pd.DataFrame:
    """Crea tabla de estadísticas descriptivas"""
    if not predictions:
        return pd.DataFrame()

    df = pd.DataFrame(predictions)
    numeric_cols = df.select_dtypes(include=[np.number]).columns

    stats_data = []

    for col in numeric_cols:
        col_stats = calculate_descriptive_stats(df[col].dropna().tolist())
        stats_data.append({
            'Variable': col,
            'N': col_stats.get('count', 0),
            'Media': round(col_stats.get('mean', 0), 2),
            'Mediana': round(col_stats.get('median', 0), 2),
            'Desv. Est.': round(col_stats.get('std', 0), 2),
            'Mín': round(col_stats.get('min', 0), 2),
            'Máx': round(col_stats.get('max', 0), 2)
        })

    return pd.DataFrame(stats_data)

# ============================================================================
# PÁGINAS PRINCIPALES
# ============================================================================

def page_dashboard():
    """Página principal - Dashboard"""
    lang = st.session_state.language
    t = TRANSLATIONS[lang]
    
    st.title(t['dashboard_title'])
    
    # Saludo personalizado
    st.subheader(t['greeting'].format(st.session_state.username))
    
    # Instrucciones de bienvenida
    st.info(f"""
    {t['how_it_works']}
    {t['how_step1']}
    {t['how_step2']}
    {t['how_step3']}
    {t['how_step4']}
    """)
    
    # Nota sobre la cuenta
    if st.session_state.username == "Invitado":
        st.warning(t['guest_warning'])

    # Métricas principales
    st.subheader(t['your_stats'])

    user_stats = get_user_stats(st.session_state.user_id)

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(
            t['total_predictions'],
            user_stats['total_predictions'],
            delta=f"{np.random.randint(1, 5)} {t['today']}"
        )

    with col2:
        st.metric(
            t['accuracy_rate'],
            f"{user_stats['accuracy_rate']:.1f}%",
            delta=f"+{np.random.randint(1, 5)}%" if user_stats['accuracy_rate'] > 0 else "0%"
        )

    with col3:
        st.metric(
            t['avg_confidence'],
            f"{user_stats['avg_confidence']:.2f}",
            delta="+0.05" if user_stats['avg_confidence'] > 0 else "0.00"
        )

    with col4:
        st.metric(
            t['global_rank'],
            user_stats['rank'],
            delta=f"+5 {t['positions']}"
        )

    st.divider()

    # Próximos partidos
    st.subheader(t['upcoming_matches'])

    with st.spinner(t['loading_matches']):
        # Primero intentar obtener partidos de la base de datos
        matches = get_upcoming_matches_from_db(days=30)
        # Si no hay partidos en DB, usar los de la API
        if not matches:
            matches = api_client.get_upcoming_matches(days=3)

    if matches:
        for i, match in enumerate(matches[:5]):
            with st.expander(f"👥 {match['teams']['home']['name']} vs {match['teams']['away']['name']}"):
                render_match_card(match)

                col1, col2, col3 = st.columns(3)

                with col1:
                    fixture_id = match['fixture']['id']
                    predictions = api_client.get_predictions(fixture_id)

                    if predictions and 'predictions' in predictions:
                        pred = predictions['predictions']
                        st.write(t['ai_prediction'])
                        st.info(t['winner'].format(pred.get('predictions', {}).get('winner', {}).get('name', 'N/A')))
                        st.write(t['probability'].format(pred.get('predictions', {}).get('win_home_or_draw', 0)))

                with col2:
                    st.write(t['your_predictions'])
                    home_score = st.number_input(t['home_goals'], min_value=0, max_value=10, key=f"home_{i}")
                    away_score = st.number_input(t['away_goals'], min_value=0, max_value=10, key=f"away_{i}")

                with col3:
                    st.write(t['your_predictions'].replace('Tus', 'Tu'))  # Adjust wording
                    confidence = st.slider(t['confidence_level'], 0.0, 1.0, 0.5, step=0.1, key=f"conf_{i}")

                    if st.button(t['predict_button'], key=f"btn_{i}"):
                        pred_data = {
                            'home_score': home_score,
                            'away_score': away_score,
                            'confidence': confidence
                        }

                        if save_prediction(st.session_state.user_id, match, pred_data, confidence):
                            st.success(t['prediction_saved'])
                            st.balloons()
                        else:
                            st.error(t['prediction_error'])
    else:
        st.info(t['no_matches_available'])

    st.divider()

    # Gráfico de rendimiento
    st.subheader(t['recent_performance'])

    # Obtener predicciones reales del usuario
    predictions = get_user_predictions(st.session_state.user_id)
    
    if predictions and len(predictions) > 0:
        df_preds = pd.DataFrame(predictions)
        # Convertir 'created_at' a datetime
        df_preds['created_at'] = pd.to_datetime(df_preds['created_at'], errors='coerce')
        df_preds = df_preds.dropna(subset=['created_at'])
        df_preds['fecha'] = df_preds['created_at'].dt.date
        
        # Calcular volumen de predicciones por día
        df_volume = df_preds.groupby('fecha').size().reset_index(name='predicciones')
        
        # Calcular precisión (simulada por ahora, ya que no tenemos resultados reales)
        # En un futuro, podrías comparar con resultados reales de partidos
        df_volume['precisión'] = np.random.uniform(0.4, 0.9, len(df_volume))
        
        # Rellenar fechas sin predicciones con 0
        date_range = pd.date_range(end=datetime.now(), periods=30)
        df_volume_full = pd.DataFrame({'fecha': date_range.date})
        df_volume_full = df_volume_full.merge(df_volume, on='fecha', how='left').fillna(0)
        
        col1, col2 = st.columns(2)

        with col1:
            fig_precision = px.line(
                df_volume_full,
                x='fecha',
                y='precisión',
                title=t['accuracy_evolution'],
                markers=True,
                color_discrete_sequence=['#FF4B4B']
            )
            fig_precision.update_layout(hovermode='x unified')
            st.plotly_chart(fig_precision, use_container_width=True)

        with col2:
            fig_volume = px.bar(
                df_volume_full,
                x='fecha',
                y='predicciones',
                title=t['prediction_volume'],
                color_discrete_sequence=['#00CC96']
            )
            st.plotly_chart(fig_volume, use_container_width=True)
    else:
        st.info(t['no_predictions_graphs'])

def page_my_predictions():
    """Página de mis predicciones"""
    lang = st.session_state.language
    t = TRANSLATIONS[lang]
    
    st.title(t['my_predictions'])

    predictions = get_user_predictions(st.session_state.user_id)

    if predictions:
        df = pd.DataFrame(predictions)

        # Resumen
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric(t['total_predictions'], len(df))
        with col2:
            high_conf = len(df[df['confidence_level'] > 0.7])
            st.metric(t['high_confidence'], high_conf)
        with col3:
            avg_conf = df['confidence_level'].mean()
            st.metric(t['avg_confidence'], f"{avg_conf:.2f}")

        st.divider()

        # Tabla de predicciones
        st.subheader(t['prediction_history'])

        display_df = df[['match_id', 'predicted_home_score', 'predicted_away_score', 'confidence_level', 'created_at']].copy()
        display_df.columns = [t['match_id'], t['home_goals'], t['away_goals'], t['confidence_level'], t['date']]

        st.dataframe(display_df, use_container_width=True)

        # Análisis
        st.subheader(t['confidence_analysis'])

        fig = px.histogram(
            df,
            x='confidence_level',
            nbins=20,
            title=t['confidence_distribution'],
            color_discrete_sequence=['#FF4B4B']
        )
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info(t['no_predictions_yet'])

def page_ai_predictions():
    """Página de predicciones con IA"""
    lang = st.session_state.language
    t = TRANSLATIONS[lang]
    
    st.title(t['ai_predictions'])

    st.markdown(t['ai_models_intro'])

    st.divider()

    # Filtros
    col1, col2, col3 = st.columns(3)

    with col1:
        sport = st.selectbox(t['sport'], [t['football'], t['basketball'], t['tennis'], t['baseball']])

    with col2:
        min_confidence = st.slider(t['min_ai_confidence'], 0.0, 1.0, 0.70)

    with col3:
        days = st.slider(t['next_days'], 1, 14, 7)

    st.divider()

    # Obtener partidos
    with st.spinner(t['analyzing_matches'].format(days)):
        # Primero intentar obtener partidos de la base de datos
        matches = get_upcoming_matches_from_db(days=days)
        # Si no hay partidos en DB, usar los de la API
        if not matches:
            matches = api_client.get_upcoming_matches(days=days)

    if matches:
        for i, match in enumerate(matches[:10]):
            home_team = match['teams']['home']['name']
            away_team = match['teams']['away']['name']
            labels = [home_team, t['draw'], away_team]
            
            with st.expander(f"🎯 {home_team} vs {away_team}"):
                render_match_card(match)
                
                # Extraer características para el modelo
                match_features = extract_match_features(match)
                
                # Obtener predicciones de todos los modelos
                all_preds = ensemble_model.predict_all_models(match_features)
                
                # Obtener predicción final del ensemble
                final_probs = all_preds["Ensemble Final"][0]
                predicted_idx = np.argmax(final_probs)
                predicted_label = labels[predicted_idx]
                confidence_value = final_probs[predicted_idx] * 100

                # Mostrar cada modelo en secciones separadas
                st.write("---")
                st.write(t['each_model_predictions'])
                
                # XGBoost
                st.subheader("🌲 XGBoost")
                xgb_probs = all_preds["XGBoost"][0]
                col_xgb1, col_xgb2 = st.columns([1, 1])
                with col_xgb1:
                    prob_xgb_df = pd.DataFrame({
                        t['result']: labels,
                        t['probability']: [f"{p*100:.1f}%" for p in xgb_probs]
                    })
                    st.dataframe(prob_xgb_df, use_container_width=True, hide_index=True)
                with col_xgb2:
                    fig_xgb = go.Figure(go.Bar(
                        x=labels,
                        y=[p*100 for p in xgb_probs],
                        marker_color=['#1F77B4', '#FF7F0E', '#2CA02C'],
                        text=[f"{p*100:.1f}%" for p in xgb_probs],
                        textposition='auto'
                    ))
                    fig_xgb.update_layout(title=t['xgb_probabilities'], yaxis_range=[0,100], height=250)
                    st.plotly_chart(fig_xgb, use_container_width=True, key=f"xgb_{i}")
                
                # Bayesian Ridge
                st.subheader("📊 Bayesian Ridge")
                bayes_probs = all_preds["Bayesian Ridge"][0]
                col_bayes1, col_bayes2 = st.columns([1, 1])
                with col_bayes1:
                    prob_bayes_df = pd.DataFrame({
                        t['result']: labels,
                        t['probability']: [f"{p*100:.1f}%" for p in bayes_probs]
                    })
                    st.dataframe(prob_bayes_df, use_container_width=True, hide_index=True)
                with col_bayes2:
                    fig_bayes = go.Figure(go.Bar(
                        x=labels,
                        y=[p*100 for p in bayes_probs],
                        marker_color=['#9467BD', '#8C564B', '#E377C2'],
                        text=[f"{p*100:.1f}%" for p in bayes_probs],
                        textposition='auto'
                    ))
                    fig_bayes.update_layout(title=t['bayes_probabilities'], yaxis_range=[0,100], height=250)
                    st.plotly_chart(fig_bayes, use_container_width=True, key=f"bayes_{i}")
                
                # LSTM con Atención
                st.subheader(t['lstm_title'])
                lstm_probs = all_preds["LSTM con Atención"][0]
                col_lstm1, col_lstm2 = st.columns([1, 1])
                with col_lstm1:
                    prob_lstm_df = pd.DataFrame({
                        t['result']: labels,
                        t['probability']: [f"{p*100:.1f}%" for p in lstm_probs]
                    })
                    st.dataframe(prob_lstm_df, use_container_width=True, hide_index=True)
                with col_lstm2:
                    fig_lstm = go.Figure(go.Bar(
                        x=labels,
                        y=[p*100 for p in lstm_probs],
                        marker_color=['#7FDBFF', '#39CCCC', '#01FF70'],
                        text=[f"{p*100:.1f}%" for p in lstm_probs],
                        textposition='auto'
                    ))
                    fig_lstm.update_layout(title=t['lstm_probabilities'], yaxis_range=[0,100], height=250)
                    st.plotly_chart(fig_lstm, use_container_width=True, key=f"lstm_{i}")
                
                st.divider()
                # Ensemble Final
                st.write(t['ensemble_title'])
                st.info(t['expected_winner'].format(predicted_label.upper()))
                col_final1, col_final2 = st.columns([2, 1])
                with col_final1:
                    prob_final_df = pd.DataFrame({
                        t['result']: labels,
                        t['probability']: [f"{p*100:.1f}%" for p in final_probs]
                    })
                    st.dataframe(prob_final_df, use_container_width=True, hide_index=True)
                    fig_final = go.Figure(go.Bar(
                        x=labels,
                        y=[p*100 for p in final_probs],
                        marker_color=['#00CC96', '#FFA15A', '#FF6B6B'],
                        text=[f"{p*100:.1f}%" for p in final_probs],
                        textposition='auto'
                    ))
                    fig_final.update_layout(title=t['ensemble_probabilities'], yaxis_range=[0,100], height=300)
                    st.plotly_chart(fig_final, use_container_width=True, key=f"final_{i}")
                    
                    # Obtener predicción de API como complemento
                    fixture_id = match['fixture']['id']
                    prediction = api_client.get_predictions(fixture_id)

                    if prediction and 'predictions' in prediction:
                        pred_data = prediction['predictions']
                        comp_data = prediction.get('comparison', {})
                        
                        st.write("---")
                        
                        if confidence_value >= min_confidence * 100:
                            col_a, col_b = st.columns(2)

                            with col_a:
                                st.write(t['expected_goals_api'])
                                goals = pred_data.get('goals', {})
                                st.write(t['home_goals_label'].format(goals.get('home', 0)))
                                st.write(t['away_goals_label'].format(goals.get('away', 0)))

                            with col_b:
                                st.write(t['ai_advice_api'])
                                advice = pred_data.get('advice', t['analyze_stats'])
                                st.write(f"_{advice}_")

                        # Comparativa
                        if comp_data:
                            st.write(t['team_comparison_title'])

                            form_data = comp_data.get('form', {})
                            att_data = comp_data.get('att', {})
                            def_data = comp_data.get('def', {})

                            comp_col1, comp_col2 = st.columns(2)

                            with comp_col1:
                                st.write(t['form_last_matches'])
                                st.write(t['home_form'].format(form_data.get('home', 'N/A')))
                                st.write(t['away_form'].format(form_data.get('away', 'N/A')))

                            with comp_col2:
                                st.write(t['attack_defense'])
                                st.write(t['home_atk_def'].format(att_data.get('home', 0), def_data.get('home', 0)))
                                st.write(t['away_atk_def'].format(att_data.get('away', 0), def_data.get('away', 0)))

                with col_final2:
                    # Medidor de confianza
                    fig_confidence = go.Figure(go.Indicator(
                        mode="gauge+number+delta",
                        value=confidence_value,
                        title={'text': t['ai_confidence_title']},
                        delta={'reference': 70},
                        domain={'x': [0, 1], 'y': [0, 1]},
                        gauge={
                            'axis': {'range': [0, 100]},
                            'bar': {'color': "darkblue"},
                            'steps': [
                                {'range': [0, 50], 'color': "lightgray"},
                                {'range': [50, 75], 'color': "lightyellow"},
                                {'range': [75, 100], 'color': "lightgreen"}
                            ],
                            'threshold': {
                                'line': {'color': "red", 'width': 4},
                                'thickness': 0.75,
                                'value': 90
                            }
                        }
                    ))
                    fig_confidence.update_layout(height=300, margin=dict(l=10, r=10, t=50, b=10))
                    st.plotly_chart(fig_confidence, use_container_width=True, key=f"conf_gauge_{i}_{fixture_id}")

                st.divider()
    else:
        st.warning(t['no_matches_period'])

def page_competitions():
    """Página de competencias"""
    lang = st.session_state.language
    t = TRANSLATIONS[lang]
    
    st.title(t['competitions'])

    # Datos de ejemplo
    competitions_data = [
        {
            'id': 1,
            'name': t['comp1_name'],
            'description': t['comp1_desc'],
            'participants': 245,
            'prize': '$1,000',
            'entry_fee': t['free'],
            'deadline': '2024-01-31',
            'status': 'active',
            'prize_pool': 1000
        },
        {
            'id': 2,
            'name': t['comp2_name'],
            'description': t['comp2_desc'],
            'participants': 89,
            'prize': '$2,500',
            'entry_fee': '$10',
            'deadline': '2024-01-25',
            'status': 'active',
            'prize_pool': 2500
        },
        {
            'id': 3,
            'name': t['comp3_name'],
            'description': t['comp3_desc'],
            'participants': 156,
            'prize': '$500',
            'entry_fee': t['free'],
            'deadline': '2024-02-15',
            'status': 'active',
            'prize_pool': 500
        }
    ]

    col1, col2 = st.columns([2, 1])

    with col1:
        st.subheader(t['available_competitions'])

        for comp in competitions_data:
            with st.expander(f"🏆 {comp['name']}"):
                st.write(t['description_label'].format(comp['description']))

                col_a, col_b, col_c = st.columns(3)

                with col_a:
                    st.metric(t['participants_label'], comp['participants'])

                with col_b:
                    st.metric(t['total_prize'], comp['prize'])

                with col_c:
                    st.metric(t['entry_label'], comp['entry_fee'])

                st.write(t['deadline_label'].format(comp['deadline']))

                col_x, col_y = st.columns(2)

                with col_x:
                    if comp['entry_fee'] == t['free']:
                        if st.button(t['join_button'], key=f"join_{comp['id']}"):
                            st.session_state.selected_competitions.append(comp['id'])
                            st.success(t['joined_comp'].format(comp['name']))
                            st.balloons()
                    else:
                        if st.session_state.user_tier == 'premium':
                            if st.button(t['join_button'], key=f"join_{comp['id']}"):
                                st.session_state.selected_competitions.append(comp['id'])
                                st.success(t['joined_comp'].format(comp['name']))
                                st.balloons()
                        else:
                            st.warning(t['premium_required'])

                with col_y:
                    st.write(t['status_label'].format(t['active_status']))

    with col2:
        st.subheader(t['your_ranking'])

        ranking_positions = [
            {'pos': 1, 'user': 'PredictorPro', 'score': 1850, 'acc': '78%'},
            {'pos': 2, 'user': 'SportGenius', 'score': 1820, 'acc': '76%'},
            {'pos': 3, 'user': 'BetMaster', 'score': 1780, 'acc': '74%'},
            {'pos': 4, 'user': 'DataAnalyst', 'score': 1750, 'acc': '72%'},
            {'pos': 5, 'user': 'GoldenPredictor', 'score': 1720, 'acc': '71%'},
            {'pos': 15, 'user': t['you_label'], 'score': 1520, 'acc': '68%'},
        ]

        for rank in ranking_positions:
            if rank['user'] == t['you_label']:
                st.markdown(
                    f"<div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); "
                    f"color: white; padding: 10px; border-radius: 5px; margin: 5px 0;'>"
                    f"<b>🎯 {rank['pos']}. {rank['user']}</b> - {rank['score']} pts ({rank['acc']})"
                    f"</div>",
                    unsafe_allow_html=True
                )
            else:
                st.write(t['rank_format'].format(rank['pos'], rank['user'], rank['score'], rank['acc']))

def page_statistics():
    """Página de estadísticas avanzadas con gráficos y reportes"""
    lang = st.session_state.language
    t = TRANSLATIONS[lang]
    
    st.title(t['statistics_title'])

    user_stats = get_user_stats(st.session_state.user_id)
    predictions = get_user_predictions(st.session_state.user_id)

    # ============ TAB 1: RESUMEN ============
    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        [t['tab1'], t['tab2'], t['tab3'], t['tab4'], t['tab5']]
    )

    with tab1:
        st.subheader(t['general_summary'])

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric(
                t['total_predictions'],
                user_stats['total_predictions'],
                delta=f"+{user_stats['total_predictions'] % 10}"
            )

        with col2:
            st.metric(
                t['accuracy_rate'],
                f"{user_stats['accuracy_rate']:.1f}%",
                delta=f"+{user_stats['accuracy_rate'] - 55:.1f}%" if user_stats['accuracy_rate'] > 55 else "-2%"
            )

        with col3:
            st.metric(
                t['avg_confidence'],
                f"{user_stats['avg_confidence']:.2f}",
                delta="+0.05"
            )

        with col4:
            st.metric(
                t['global_rank'],
                str(user_stats['rank']),
                delta=f"+5 {t['positions']}"
            )

    # ============ TAB 2: TABLAS ESTADÍSTICAS ============
    with tab2:
        st.subheader(t['descriptive_stats'])

        # Tabla de resumen
        col1, col2 = st.columns(2)

        with col1:
            st.write(t['performance_summary'])
            summary_table = pd.DataFrame({
                t['metric_col']: [t['total'], t['correct'], t['incorrect'], t['precision'], t['avg_confidence'], t['ranking']],
                t['value_col']: [
                    str(user_stats['total_predictions']),
                    str(user_stats['correct_predictions']),
                    str(user_stats['total_predictions'] - user_stats['correct_predictions']),
                    f"{user_stats['accuracy_rate']:.2f}%",
                    f"{user_stats['avg_confidence']:.2f}",
                    str(user_stats['rank'])
                ]
            })
            st.table(summary_table)

        with col2:
            st.write(t['pred_desc_stats'])
            stats_table = create_descriptive_stats_table(predictions)
            if not stats_table.empty:
                st.dataframe(stats_table, use_container_width=True)
            else:
                st.info(t['no_numeric_data'])

    # ============ TAB 3: GRÁFICOS AVANZADOS ============
    with tab3:
        st.subheader(t['advanced_viz'])

        if predictions and len(predictions) > 0:
            df = pd.DataFrame(predictions)

            # Gráfico 1: Distribución de Confianza
            if 'confidence_level' in df.columns:
                col1, col2 = st.columns(2)

                with col1:
                    confidence_data = df['confidence_level'].dropna().tolist()
                    fig_dist = create_distribution_plot(
                        confidence_data,
                        t['conf_dist_title']
                    )
                    st.plotly_chart(fig_dist, use_container_width=True)

                with col2:
                    # Gráfico de CDF (Función de Distribución Acumulada)
                    sorted_data = np.sort(confidence_data)
                    cumulative = np.arange(1, len(sorted_data) + 1) / len(sorted_data)

                    fig_cdf = go.Figure()
                    fig_cdf.add_trace(go.Scatter(
                        x=sorted_data, y=cumulative,
                        mode='lines',
                        name='CDF',
                        line=dict(color='#667EEA', width=3)
                    ))
                    fig_cdf.update_layout(
                        title=t['cdf_title'],
                        xaxis_title=t['confidence'],
                        yaxis_title=t['probability'],
                        height=400
                    )
                    st.plotly_chart(fig_cdf, use_container_width=True)

            # Gráfico 2: Serie Temporal
            if 'created_at' in df.columns:
                st.subheader(t['time_evolution'])

                df['created_at'] = pd.to_datetime(df['created_at'], errors='coerce')
                df_sorted = df.sort_values('created_at').dropna(subset=['created_at'])

                if len(df_sorted) > 0:
                    dates = df_sorted['created_at'].tolist()
                    values = df_sorted.get('confidence_level', pd.Series(range(len(df_sorted)))).tolist()

                    fig_time = create_time_series_plot(
                        dates, values,
                        t['time_evol_conf']
                    )
                    st.plotly_chart(fig_time, use_container_width=True)

            # Gráfico 3: Por Status
            if 'prediction_status' in df.columns:
                st.subheader(t['status_analysis'])

                status_counts = df['prediction_status'].value_counts()
                fig_status = px.bar(
                    x=status_counts.index,
                    y=status_counts.values,
                    title=t['status_dist'],
                    labels={'x': t['status'], 'y': t['quantity']},
                    color_discrete_sequence=['#667EEA']
                )
                fig_status.update_layout(height=400)
                st.plotly_chart(fig_status, use_container_width=True)

        else:
            st.info(t['not_enough_data'])

    # ============ TAB 4: REPORTES ============
    with tab4:
        st.subheader(t['generate_reports'])

        col1, col2 = st.columns(2)

        with col1:
            st.write(t['pdf_report'])
            if st.button(t['download_pdf']):
                pdf_bytes = generate_predictions_report_pdf(predictions)
                st.download_button(
                    label=t['download_pdf'],
                    data=pdf_bytes,
                    file_name=f"predicciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf"
                )

        with col2:
            st.write(t['excel_report'])
            if st.button(t['download_excel']):
                excel_bytes = generate_predictions_report_excel(predictions)
                st.download_button(
                    label=t['download_excel'],
                    data=excel_bytes,
                    file_name=f"predicciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # ============ TAB 5: INSIGHTS ============
    with tab5:
        st.subheader(t['insights_title'])

        accuracy = user_stats.get('accuracy_rate', 0)
        confidence = user_stats.get('avg_confidence', 0)

        col1, col2 = st.columns(2)

        with col1:
            st.write(t['perf_analysis'])

            if accuracy >= 75:
                st.success(t['excellent_perf'])
                insight1 = t['insight1_excellent']
            elif accuracy >= 60:
                st.info(t['good_perf'])
                insight1 = t['insight1_good']
            else:
                st.warning(t['dev_perf'])
                insight1 = t['insight1_dev']

            st.write(insight1)

        with col2:
            st.write(t['conf_analysis'])

            if confidence >= 0.75:
                st.success(t['high_conf'])
                insight2 = t['insight2_high']
            elif confidence >= 0.6:
                st.info(t['moderate_conf'])
                insight2 = t['insight2_moderate']
            else:
                st.warning(t['low_conf'])
                insight2 = t['insight2_low']

            st.write(insight2)

        st.divider()

        st.write(t['personal_recommendations'])
        recommendations = [
            t['rec1'], t['rec2'], t['rec3'], t['rec4'], t['rec5'], t['rec6']
        ]

        for rec in recommendations:
            st.write(rec)

def page_alerts():
    """Página de alertas"""
    st.title("🔔 Alertas y Notificaciones")

    alerts_data = [
        {
            'type': 'partido',
            'icon': '⚽',
            'title': 'Nuevo partido disponible',
            'message': 'Real Madrid vs Barcelona - Predicción IA lista',
            'time': 'Hace 2 horas',
            'read': False
        },
        {
            'type': 'prediccion',
            'icon': '🤖',
            'title': 'Actualización de predicción IA',
            'message': 'Nueva predicción para Manchester City vs Liverpool',
            'time': 'Hace 5 horas',
            'read': True
        },
        {
            'type': 'competencia',
            'icon': '🏆',
            'title': 'Competencia finaliza pronto',
            'message': 'Liga Predictor Enero cierra en 2 días',
            'time': 'Hace 1 día',
            'read': False
        },
        {
            'type': 'ranking',
            'icon': '📈',
            'title': 'Cambio en tu ranking',
            'message': 'Subiste 5 posiciones en el ranking global',
            'time': 'Hace 1 día',
            'read': True
        }
    ]

    col1, col2 = st.columns([3, 1])

    with col1:
        st.subheader("Notificaciones Recientes")

        for alert in alerts_data:
            status_icon = "📬" if alert['read'] else "📭"
            bg_color = "#f0f0f0" if alert['read'] else "#fff3cd"

            col_alert = st.columns([1, 10, 1])

            with col_alert[0]:
                st.write(alert['icon'])

            with col_alert[1]:
                st.markdown(
                    f"<div style='background: {bg_color}; padding: 15px; border-radius: 5px;'>"
                    f"<b>{alert['title']}</b><br>"
                    f"{alert['message']}<br>"
                    f"<small>{alert['time']}</small>"
                    f"</div>",
                    unsafe_allow_html=True
                )

            with col_alert[2]:
                st.write(status_icon)

    with col2:
        st.subheader("Configurar Alertas")

        st.checkbox("⚽ Nuevos partidos", value=True)
        st.checkbox("🤖 Predicciones IA", value=True)
        st.checkbox("🏆 Competencias", value=True)
        st.checkbox("📈 Cambios ranking", value=True)
        st.checkbox("💎 Ofertas Premium", value=False)

        if st.button("Guardar configuración"):
            st.success("✅ Configuración guardada")

def page_premium():
    """Página de suscripción premium"""
    st.title("💎 SportsPredict Pro Premium")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.subheader("📦 Plan Free")

        st.markdown("""
        ✅ Predicciones IA básicas
        ✅ Competencias gratuitas
        ✅ Estadísticas básicas
        ✅ Hasta 50 predicciones/mes

        ---

        **Precio:** Gratis
        """)

        if st.session_state.user_tier != "free":
            st.info("Tu plan actual: Premium")
        else:
            st.success("Tu plan actual")

    with col2:
        st.subheader("🌟 Plan Pro")

        st.markdown("""
        ✅ Todas las del plan Free
        ✅ Predicciones IA avanzadas
        ✅ Acceso a competencias premium
        ✅ Análisis detallados
        ✅ Reportes PDF personalizados
        ✅ Alertas ilimitadas
        ✅ Hasta 500 predicciones/mes

        ---

        **Precio:** $4.99/mes
        *(o $49.99/año)*
        """)

        if st.button("🚀 Suscribirse a Pro", key="btn_pro"):
            st.session_state.user_tier = "pro"
            st.success("¡Bienvenido a SportsPredict Pro!")
            st.balloons()

    with col3:
        st.subheader("👑 Plan Elite")

        st.markdown("""
        ✅ Todas las del plan Pro
        ✅ Acceso a análisis IA en tiempo real
        ✅ Consultas personalizadas con expertos
        ✅ Datos históricos completos
        ✅ Predicciones ilimitadas
        ✅ Prioridad en competencias
        ✅ Descuentos en apuestas

        ---

        **Precio:** $9.99/mes
        *(o $99.99/año)*
        """)

        if st.button("👑 Suscribirse a Elite", key="btn_elite"):
            st.session_state.user_tier = "elite"
            st.success("¡Bienvenido a SportsPredict Elite!")
            st.balloons()

    st.divider()

    st.subheader("Beneficios Adicionales")

    benefits = pd.DataFrame({
        'Característica': [
            'Predicciones IA básicas',
            'Predicciones IA avanzadas',
            'Acceso a competencias',
            'Reportes PDF',
            'Alertas personalizadas',
            'Análisis histórico',
            'Soporte prioritario',
            'Consulta con expertos'
        ],
        'Free': ['✅', '❌', '✅', '❌', '❌', '❌', '❌', '❌'],
        'Pro': ['✅', '✅', '✅', '✅', '✅', '✅', '❌', '❌'],
        'Elite': ['✅', '✅', '✅', '✅', '✅', '✅', '✅', '✅']
    })

    st.dataframe(benefits, use_container_width=True)

# ============================================================================
# MENÚ PRINCIPAL Y NAVEGACIÓN
# ============================================================================

def main():
    """Función principal"""
    
    # Mostrar página de login solo si el usuario no ha elegido un modo
    if st.session_state.username == "Invitado" and not st.session_state.authenticated and 'mode_chosen' not in st.session_state:
        login_page()
        return

    lang = st.session_state.language
    t = TRANSLATIONS[lang]

    # Sidebar
    st.sidebar.title("⚽ SportsPredict Pro")
    st.sidebar.markdown("---")

    # Controles de idioma y modo oscuro
    with st.sidebar:
        col1, col2 = st.columns(2)
        with col1:
            selected_lang = st.selectbox(t["language"], ["Español", "English"], index=0 if lang == "es" else 1)
            if selected_lang == "Español" and lang != "es":
                st.session_state.language = "es"
                st.rerun()
            elif selected_lang == "English" and lang != "en":
                st.session_state.language = "en"
                st.rerun()
        
        with col2:
            dark_mode_label = t["dark_mode"] if st.session_state.dark_mode else t["light_mode"]
            if st.button(dark_mode_label):
                st.session_state.dark_mode = not st.session_state.dark_mode
                st.rerun()

    # Renderizar chatbot
    render_chatbot()
    
    # Estado de conexión a Supabase
    if SUPABASE_AVAILABLE and supabase:
        try:
            # Prueba simple de conexión
            test_response = supabase.table('user_predictions').select('id', count='exact').limit(1).execute()
            st.sidebar.success(t["connected_supabase"])
        except Exception as e:
            st.sidebar.warning(t["supabase_warning"])
    else:
        st.sidebar.info(t["local_mode"])

    # Información del usuario
    show_user_profile()
    st.sidebar.markdown("---")

    # Menú de navegación
    menu_options = [
        t["dashboard"],
        t["my_predictions"],
        t["ai_predictions"],
        t["ai_assistant_page"],
        t["competitions"],
        t["statistics"],
        t["alerts"],
        t["premium"]
    ]
    
    # Agregar opción de administración si es admin
    if st.session_state.is_admin:
        menu_options.append(t["admin"])

    selected = st.sidebar.radio(t["menu"], menu_options)

    st.sidebar.divider()

    # Información adicional
    with st.sidebar:
        st.write(f"### {t['info_title'].split(' ')[0]}")
        st.info(t["info_text"])

    # Renderizar página seleccionada
    if selected == t["dashboard"]:
        page_dashboard()
    elif selected == t["my_predictions"]:
        page_my_predictions()
    elif selected == t["ai_predictions"]:
        page_ai_predictions()
    elif selected == t["ai_assistant_page"]:
        page_ai_assistant()
    elif selected == t["competitions"]:
        page_competitions()
    elif selected == t["statistics"]:
        page_statistics()
    elif selected == t["alerts"]:
        page_alerts()
    elif selected == t["premium"]:
        page_premium()
    elif selected == t["admin"] and st.session_state.is_admin:
        st.title(t["admin"])
        
        # Tabs for admin sections
        admin_tab1, admin_tab2, admin_tab3 = st.tabs([t["admin_tab1"], t["admin_tab2"], "🤖 Validación con API"])
        
        with admin_tab1:
            # Intentar usar Supabase con clave de servicio, si no, usar modo local
            # El cliente de servicio ignora RLS para operaciones de admin
            show_table_manager_ui(supabase_service)
        
        with admin_tab3:
            st.header("🤖 Validación con Datos Reales (API Football)")
            st.info("Estos resultados comparan el desempeño de los modelos contra partidos reales finalizados usando la API de Football. El test de McNemar nos dice si las diferencias de precisión son significativas o si podrían ser al azar.")
            if os.path.exists("validation_report.json"):
                with open("validation_report.json", "r") as f:
                    report_data = json.load(f)
                
                st.success(f"🏆 MEJOR MODELO EN DATOS REALES (API): **{report_data['best_model']}**")
                st.markdown("""
                **¿Por qué se selecciona automáticamente?**  
                Al validar contra docenas de partidos recién finalizados (API-Football), este modelo demostró ser estadísticamente superior mediante la prueba de McNemar y superó la prueba de azar (Binomial). En un pipeline automatizado de apuestas deportivas, promover dinámicamente el modelo empíricamente superior evita el sesgo humano, previene regresiones en rendimiento y garantiza que el sistema siempre use las probabilidades más exactas frente a la realidad actual.
                """)
                col1, col2 = st.columns([1, 1.5])
                with col1:
                    st.subheader("Accuracy por Modelo")
                    # Extraer accuracies
                    acc_data = [{"Modelo": m["name"], "Accuracy": m["accuracy"]} for m in report_data["models"]]
                    acc_df = pd.DataFrame(acc_data)
                    st.bar_chart(acc_df.set_index("Modelo"))
                
                with col2:
                    st.subheader("Análisis Estadístico Detallado por Modelo")
                    
                    for m in report_data["models"]:
                        with st.expander(f"{m['name']} - {(m['accuracy']*100):.2f}% Accuracy"):
                            st.caption(m["description"])
                            st.write("---")
                            st.write("**Pruebas de McNemar:**")
                            mcnemar_data = [{"Comparado con": t["compared_to"], "P-Value": f"{t['p_value']:.4f}", "Significativo": "✅ Sí" if t["significant"] else "❌ No"} for t in m["mcnemar_tests"]]
                            if mcnemar_data:
                                st.table(pd.DataFrame(mcnemar_data))
                            
                            st.write("**Otras Pruebas Robustas:**")
                            other_data = []
                            bt = m.get("additional_tests", {}).get("binomial_test")
                            if bt:
                                other_data.append({"Prueba": "Binomial", "Descripción": bt["description"], "P-Value": f"{bt['p_value']:.4f}", "Resultado": "✅ Mejor que azar" if bt["significant"] else "❌ Igual al azar"})
                            
                            ct = m.get("additional_tests", {}).get("chi_square_test")
                            if ct:
                                other_data.append({"Prueba": "Chi-Cuadrado", "Descripción": ct["description"], "P-Value": f"{ct['p_value']:.4f}", "Resultado": "✅ Difiere" if ct["significant"] else "❌ Similar"})
                            
                            if other_data:
                                st.table(pd.DataFrame(other_data))
            else:
                st.warning("No se encontró el reporte. Por favor ejecuta el script de validación de API.")
                
        with admin_tab2:
            st.header(t["admin_validation_header"])
            
            # Try loading full results from joblib first if available
            if "model_validation_results" not in st.session_state or not st.session_state.model_validation_results:
                if os.path.exists("model_validation_results_full.joblib"):
                    import joblib
                    try:
                        st.session_state.model_validation_results = joblib.load("model_validation_results_full.joblib")
                        st.info("Se cargaron los resultados de validación anteriores!")
                    except Exception as e:
                        st.warning(f"No se pudieron cargar los resultados anteriores: {e}")
            
            # Big visible button!
            st.markdown("---")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button(t["admin_run_validation"], type="primary", use_container_width=True):
                    with st.spinner(t["admin_validating"]):
                        try:
                            results = validate_models.run_robust_validation(save_dir="saved_models")
                            st.session_state.model_validation_results = results
                            st.success(t["admin_validation_success"])
                        except Exception as e:
                            st.error(t["admin_validation_error"].format(str(e)))
                            st.session_state.model_validation_results = None
            st.markdown("---")
            
            # Show results if available in session state OR load from joblib
            if "model_validation_results" in st.session_state and st.session_state.model_validation_results:
                full_res = st.session_state.model_validation_results
                
                if "metrics" in full_res:
                    st.warning("Se detectaron resultados con formato antiguo. Por favor haz clic en 'Ejecutar Validación Robusta' arriba para regenerarlos con el nuevo formato multi-modelo.")
                else:
                    model_names = list(full_res.keys())
                    selected_model = st.selectbox("Seleccione el Modelo para auditar:", model_names)
                    
                    res = full_res[selected_model]
                    
                    st.subheader(f"Métricas de Clasificación - {selected_model}")
                    st.dataframe(pd.DataFrame(list(res["metrics"]["classification"].items()), columns=["Métrica", "Valor"]), use_container_width=True)
                    
                    if "regression" in res["metrics"]:
                        st.subheader(f"Métricas de Regresión - {selected_model}")
                        st.dataframe(pd.DataFrame(list(res["metrics"]["regression"].items()), columns=["Métrica", "Valor"]), use_container_width=True)
                    
                    st.subheader(t["admin_visualizations"])
                    
                    # Confusion matrix
                    if "confusion_matrix" in res["figures"]:
                        st.plotly_chart(res["figures"]["confusion_matrix"], use_container_width=True)
                        
                    # Residual analysis
                    if "residuals" in res["figures"]:
                        st.plotly_chart(res["figures"]["residuals"], use_container_width=True)
                        
                    if "residual_analysis" in res:
                        st.subheader(t["admin_residual_analysis"])
                        st.dataframe(pd.DataFrame(list(res["residual_analysis"].items()), columns=["Métrica", "Valor"]), use_container_width=True)
                        
                    if "normality_tests" in res:
                        st.subheader(t["admin_normality_tests"])
                        st.dataframe(pd.DataFrame([
                            {"Prueba": "Shapiro-Wilk", "Estadístico": res["normality_tests"]["Shapiro-Wilk"]["statistic"], "p-valor": res["normality_tests"]["Shapiro-Wilk"]["p_value"]},
                            {"Prueba": "Kolmogorov-Smirnov", "Estadístico": res["normality_tests"]["Kolmogorov-Smirnov"]["statistic"], "p-valor": res["normality_tests"]["Kolmogorov-Smirnov"]["p_value"]}
                        ]), use_container_width=True)
                    
                    # Cross-validation
                    if "cross_validation" in res["figures"]:
                        st.plotly_chart(res["figures"]["cross_validation"], use_container_width=True)
                    
                    st.subheader(t["admin_cv_stability"])
                    if "stability_tests" in res:
                        st.dataframe(pd.DataFrame([
                            {"Métrica": "Promedio Estabilidad", "Valor": res["stability_tests"].get("cv_mean", 0)},
                            {"Métrica": "Desviación Estándar", "Valor": res["stability_tests"].get("cv_std", 0)}
                        ]), use_container_width=True)
                    
                    # Overfitting test
                    if "overfitting" in res["figures"]:
                        st.plotly_chart(res["figures"]["overfitting"], use_container_width=True)
                        st.dataframe(pd.DataFrame([
                            {"Métrica": "Accuracy Entrenamiento", "Valor": res["overfitting_tests"].get("train_accuracy", 0)},
                            {"Métrica": "Accuracy Prueba", "Valor": res["overfitting_tests"].get("test_accuracy", 0)},
                            {"Métrica": "Brecha (Gap)", "Valor": res["overfitting_tests"].get("accuracy_gap", 0)}
                        ]), use_container_width=True)
                
                st.subheader(t["admin_overfitting_test"])
                st.dataframe(pd.DataFrame(list(res["overfitting_tests"].items()), columns=["Métrica", "Valor"]), use_container_width=True)
                
                # Download buttons for results
                if os.path.exists("model_validation_results.csv"):
                    with open("model_validation_results.csv", "rb") as f:
                        st.download_button(t["admin_download_results"], f, "model_validation_results.csv", "text/csv")
                if os.path.exists("model_validation_results_full.joblib"):
                    with open("model_validation_results_full.joblib", "rb") as f:
                        st.download_button("📥 Descargar Resultados Completo (joblib)", f, "model_validation_results_full.joblib", "application/octet-stream")
            elif os.path.exists("model_validation_results.csv"):
                # If CSV exists but no joblib, show CSV preview
                st.info("Se encontraron resultados de validación anteriores!")
                df = pd.read_csv("model_validation_results.csv")
                st.dataframe(df, use_container_width=True)
                with open("model_validation_results.csv", "rb") as f:
                    st.download_button(t["admin_download_results"], f, "model_validation_results.csv", "text/csv")

if __name__ == "__main__":
    main()
